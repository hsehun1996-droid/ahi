# -*- coding: utf-8 -*-
"""고속도로 포장유지보수 이력 관리 - CSV/Excel/PS/PDF 저장·불러오기·내보내기, SQLite 캐시
"""
import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import csv
import customtkinter as ctk
import math
import sqlite3
import json
import logging
from datetime import datetime
import tkinter.font as tkfont
import io
import ctypes
import tempfile

try:
    from PIL import Image, ImageOps, ImageTk
    from reportlab.pdfgen import canvas as reportlab_canvas
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.utils import ImageReader
    _PDF_LIBS_AVAILABLE = True
except ImportError:
    _PDF_LIBS_AVAILABLE = False
    ImageTk = None

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    _EXCEL_LIBS_AVAILABLE = True
except ImportError:
    _EXCEL_LIBS_AVAILABLE = False

try:
    from CTkMessagebox import CTkMessagebox
except ImportError:
    CTkMessagebox = None

# 개별 파일을 직접 실행하거나 작업 폴더가 달라도 프로젝트 루트(상위 폴더)의
# constants/utils 등을 찾을 수 있도록 sys.path 에 추가
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from constants import *
from utils import (
    get_logger, log_exception, log_warning,
    read_csv_with_encoding, get_user_data_dir, resource_path,
    load_method_settings, reset_method_settings,
    get_method_warranty_period, build_source_signature,
    km_floor_to_grid, km_ceil_to_grid, clamp, fmt_km,
    intervals_overlap, lane_conflict,
)
from canvas_utils import (
    _split_gaps, _merge_gaps, _draw_hatch_rect, _compute_overlap_status,
)


class IOMixin:
    def on_save_csv(self):
        # 현재 선택된 노선의 이력 CSV + IC CSV를 개별 파일로 저장 (대화상자)
        if self.current_route_index < 0 or self.current_route_index >= len(self.routes):
            self._show_error("노선 없음", "저장할 노선이 없습니다. 먼저 노선을 추가해 주세요.")
            return
        route = self.routes[self.current_route_index]
        fpath = filedialog.asksaveasfilename(
            title="CSV로 저장",
            defaultextension=".csv",
            filetypes=[("CSV 파일", "*.csv"), ("모든 파일", "*.*")],
            initialfile=f"{route['name']}_maintenance_history.csv",
        )
        if not fpath:
            return
        try:
            with open(fpath, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow([
                    "route_name", "route_start_km", "route_end_km",
                    "start_km", "end_km", "method", "direction", "lane", "work_date", "timestamp",
                    "row_type", "route_dir1", "route_dir2"
                ])
                wrote_any = False
                dirs = route.get("directions") or DIRECTIONS
                dir1 = dirs[0]
                dir2 = dirs[1 if len(dirs) > 1 else 0]
                for it in route.get("entries", []):
                    if it.get("plan"):
                        continue  # 사업계획 구간은 all_business_plan.csv에 별도 저장
                    wrote_any = True
                    writer.writerow([
                        route.get("name", ""), route.get("start_km", 0.0), route.get("end_km", 0.0),
                        it.get("start", 0.0), it.get("end", 0.0), it.get("method", ""), it.get("direction", ""), it.get("lane", ""), it.get("work_date", ""), it.get("ts", ""),
                        "entry", dir1, dir2
                    ])
                if not wrote_any:
                    writer.writerow([
                        route.get("name", ""), route.get("start_km", 0.0), route.get("end_km", 0.0),
                        "", "", "", "", "", "",
                        "meta", dir1, dir2
                    ])

            # 동일 폴더에 IC CSV도 함께 저장
            base_dir = os.path.dirname(fpath)
            ic_fname = f"{route.get('name','')}_ic.csv"
            ic_fpath = os.path.join(base_dir, ic_fname)
            ic_rows = []
            for ic in getattr(self, 'ics', []):
                ic_route = str(ic.get('route') or '')
                if ic_route and ic_route != route.get('name', ''):
                    continue
                km = float(ic.get('km', 0.0))
                name = str(ic.get('name') or '')
                ramps = ic.get('ramps') or []
                if not ramps:
                    ic_rows.append([route.get('name',''), name, km, '', ''])
                else:
                    for r in ramps:
                        if isinstance(r, dict):
                            nm = str(r.get('name') or '')
                            try:
                                rkm = float(r.get('km', km))
                            except Exception:
                                rkm = km
                            ic_rows.append([route.get('name',''), name, km, nm, rkm])
            with open(ic_fpath, "w", newline="", encoding="utf-8-sig") as f2:
                w = csv.writer(f2)
                w.writerow(["route_name", "ic_name", "ic_km", "ramp_name", "ramp_km"])
                for row in ic_rows:
                    w.writerow(row)

            # 구조물 CSV 저장
            structures = route.get('structures', [])
            if structures:
                struct_fname = f"{route.get('name','')}_structures.csv"
                struct_fpath = os.path.join(base_dir, struct_fname)
                with open(struct_fpath, "w", newline="", encoding="utf-8-sig") as f_struct:
                    writer = csv.writer(f_struct)
                    writer.writerow(["route_name", "name", "type", "direction", "start_km", "end_km"])
                    for s in structures:
                        writer.writerow([
                            route.get('name', ''), s.get('name', ''), s.get('type', ''),
                            s.get('direction', '양방향'),
                            s.get('start', 0.0), s.get('end', 0.0)
                        ])


            # 지사 이력 CSV 저장
            branch_entries = route.get("branch_entries", [])
            if branch_entries:
                branch_fname = f"{route.get('name','')}_branch_history.csv"
                branch_fpath = os.path.join(base_dir, branch_fname)
                with open(branch_fpath, "w", newline="", encoding="utf-8-sig") as f_br:
                    w_br = csv.writer(f_br)
                    w_br.writerow([
                        "route_name", "route_start_km", "route_end_km",
                        "start_km", "end_km", "method", "direction", "lane", "work_date", "timestamp"
                    ])
                    for it in branch_entries:
                        w_br.writerow([
                            route.get("name", ""), route.get("start_km", 0.0), route.get("end_km", 0.0),
                            it.get("start", 0.0), it.get("end", 0.0), it.get("method", ""),
                            it.get("direction", ""), it.get("lane", ""), it.get("work_date", ""), it.get("ts", "")
                        ])

            self._show_info("완료", "CSV 저장이 완료되었습니다.")
        except Exception as ex:
            self._show_error("오류", f"저장 실패: {ex}")

    def on_load_csv(self):
        # 현재 선택된 노선으로 불러오기 (대화상자)
        if self.current_route_index < 0 or self.current_route_index >= len(self.routes):
            self._show_error("노선 없음", "불러올 노선이 없습니다. 먼저 노선을 추가해 주세요.")
            return
        route = self.routes[self.current_route_index]
        fpath = filedialog.askopenfilename(
            title="CSV 불러오기",
            filetypes=[("CSV 파일", "*.csv"), ("모든 파일", "*.*")],
        )
        if not fpath:
            return
        loaded = []
        first_row = None
        def _read_with_encoding(enc: str):
            nonlocal loaded, first_row
            with open(fpath, "r", encoding=enc) as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if first_row is None:
                        first_row = row
                    try:
                        loaded.append({
                            "start": float(row.get("start_km") or row.get("start") or 0.0),
                            "end": float(row.get("end_km") or row.get("end") or 0.0),
                            "method": row.get("method", "절삭 덧씌우기"),
                            "direction": (row.get("direction") or "").strip(),
                            "lane": row.get("lane", "전차로"),
                            "work_date": row.get("work_date", ""),
                            "ts": row.get("timestamp", ""),
                            "row_type": row.get("row_type", "entry"),
                            "route_dir1": row.get("route_dir1", ""),
                            "route_dir2": row.get("route_dir2", ""),
                        })
                    except Exception:
                        continue

        try:
            read_csv_with_encoding(_read_with_encoding)
        except Exception as ex:
            self._show_error("오류", f"불러오기 실패: {ex}")
            return

        # 파일 내 노선 시작/끝 이정/방향 메타가 있으면 반영
        if first_row:
            try:
                rs = float(first_row.get("route_start_km") or first_row.get("start_km") or first_row.get("start") or route["start_km"]) 
                re = float(first_row.get("route_end_km") or first_row.get("end_km") or first_row.get("end") or route["end_km"]) 
                route["start_km"], route["end_km"] = rs, re
            except Exception:
                pass
            # 방향 메타 복원
            d1 = (first_row.get("route_dir1") or "").strip()
            d2 = (first_row.get("route_dir2") or "").strip()
            if d1 and d2:
                route["directions"] = [d1, d2]

        # entry만 entries로 채움
        route["entries"] = [
            {k: it[k] for k in ("start","end","method","direction","lane","work_date","ts")}
            for it in loaded if (it.get("row_type") or "entry") == "entry"
        ]
        self._mark_route_data_changed(route)
        # 불러온 이력에서 실제 사용된 방향 문자열을 반영
        try:
            # CSV 메타에 방향이 있으면 우선 사용, 없으면 이력에서 추정
            if not route.get("directions") or len(route.get("directions", [])) < 2:
                self.infer_route_directions_from_entries(route)
            # 컨트롤 갱신
            self.refresh_route_controls_from_route(route)
        except Exception:
            pass
        # 지사 이력 자동 불러오기 (동일 폴더의 {name}_branch_history.csv)
        branch_fpath = os.path.join(os.path.dirname(fpath),
                                    f"{route.get('name','')}_branch_history.csv")
        if os.path.exists(branch_fpath):
            branch_loaded = []
            def _read_branch(enc):
                with open(branch_fpath, "r", encoding=enc) as fb:
                    reader = csv.DictReader(fb)
                    for row in reader:
                        try:
                            branch_loaded.append({
                                "start": float(row.get("start_km") or 0.0),
                                "end": float(row.get("end_km") or 0.0),
                                "method": row.get("method", "단면보수"),
                                "direction": (row.get("direction") or "").strip(),
                                "lane": row.get("lane", "전차로"),
                                "work_date": row.get("work_date", ""),
                                "ts": row.get("timestamp", ""),
                            })
                        except Exception:
                            continue
            try:
                read_csv_with_encoding(_read_branch)
            except Exception:
                pass
            route["branch_entries"] = branch_loaded
            self._mark_route_data_changed(route)

        self.update_year_filter_values()
        self.draw_schematic()
        self._show_info("불러오기", f"불러오기 완료: {len(loaded)}건")

    def on_load_excel(self):
        """Excel 파일(.xlsx)에서 유지보수 이력을 불러옵니다."""
        if not _EXCEL_LIBS_AVAILABLE:
            self._show_error(
                "라이브러리 필요",
                "Excel 불러오기를 사용하려면 'openpyxl' 라이브러리가 필요합니다.\n\n"
                "터미널(cmd)에서 아래 명령어를 실행하세요:\n"
                "pip install openpyxl"
            )
            return
        if self.current_route_index < 0 or self.current_route_index >= len(self.routes):
            self._show_error("노선 없음", "불러올 노선이 없습니다. 먼저 노선을 추가해 주세요.")
            return
        route = self.routes[self.current_route_index]

        fpath = filedialog.askopenfilename(
            title="Excel 불러오기",
            filetypes=[("Excel 파일", "*.xlsx *.xlsm"), ("모든 파일", "*.*")],
        )
        if not fpath:
            return

        try:
            wb = load_workbook(fpath, data_only=True)
            ws = wb.active
        except Exception as ex:
            self._show_error("오류", f"Excel 파일을 열 수 없습니다: {ex}")
            return

        # 헤더 행 찾기 (첫 번째 비어있지 않은 행)
        headers = []
        header_row_idx = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            non_empty = [c for c in row if c is not None]
            if len(non_empty) >= 3:
                headers = [str(c).strip() if c is not None else "" for c in row]
                header_row_idx = i
                break

        if not headers:
            self._show_error("오류", "헤더 행을 찾을 수 없습니다. 파일 형식을 확인해 주세요.")
            return

        # 컬럼 자동 매핑 (CSV 컬럼명과 유사한 이름 우선 탐지)
        col_map = {}
        aliases = {
            "start": ["start_km", "start", "시작이정", "시점", "시작km", "시작"],
            "end":   ["end_km", "end", "끝이정", "종점", "끝km", "끝"],
            "method": ["method", "공법", "보수공법", "포장공법"],
            "direction": ["direction", "방향", "노선방향"],
            "lane": ["lane", "차로", "차선"],
            "work_date": ["work_date", "시공일자", "작업일자", "날짜", "일자", "date"],
        }
        lower_headers = [h.lower() for h in headers]
        for field, names in aliases.items():
            for name in names:
                if name.lower() in lower_headers:
                    col_map[field] = lower_headers.index(name.lower())
                    break

        # 필수 컬럼(start, end)이 자동 탐지 안 되면 매핑 다이얼로그 표시
        if "start" not in col_map or "end" not in col_map:
            col_map = self._show_excel_column_mapping_dialog(headers, col_map)
            if col_map is None:
                return

        loaded = []
        for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
            if all(c is None for c in row):
                continue
            try:
                def _cell(idx):
                    if idx is None:
                        return None
                    v = row[idx] if idx < len(row) else None
                    # openpyxl이 날짜 셀을 datetime으로 반환하는 경우 처리
                    if hasattr(v, "strftime"):
                        return v.strftime("%Y%m%d")
                    return str(v).strip() if v is not None else ""

                start_val = _cell(col_map.get("start"))
                end_val = _cell(col_map.get("end"))
                if not start_val and not end_val:
                    continue
                loaded.append({
                    "start": float(start_val or 0),
                    "end": float(end_val or 0),
                    "method": _cell(col_map.get("method")) or "절삭 덧씌우기",
                    "direction": _cell(col_map.get("direction")) or "",
                    "lane": _cell(col_map.get("lane")) or "전차로",
                    "work_date": _cell(col_map.get("work_date")) or "",
                    "ts": "",
                })
            except Exception:
                continue

        if not loaded:
            self._show_error("결과 없음", "불러온 이력이 없습니다. 파일 내용을 확인해 주세요.")
            return

        route["entries"] = loaded
        self._mark_route_data_changed(route)
        try:
            if not route.get("directions") or len(route.get("directions", [])) < 2:
                self.infer_route_directions_from_entries(route)
            self.refresh_route_controls_from_route(route)
        except Exception:
            pass

        self.update_year_filter_values()
        self.draw_schematic()
        self._show_info("불러오기 완료", f"Excel에서 {len(loaded)}건을 불러왔습니다.")

    def _show_excel_column_mapping_dialog(self, headers: list, pre_map: dict) -> dict | None:
        """사용자가 Excel 컬럼을 수동으로 매핑하는 다이얼로그를 표시합니다."""
        dlg = self._create_popup_window(self)
        dlg.title("")
        dlg.resizable(False, False)
        dlg.grab_set()

        ctk.CTkLabel(dlg, text="Excel 파일의 컬럼을 각 항목에 맞게 선택해 주세요.",
                     font=(self.font_family, 13)).pack(padx=20, pady=(16, 8))

        choices_with_blank = ["(없음)"] + headers
        fields = [
            ("start",     "시작 이정(km) *"),
            ("end",       "끝 이정(km) *"),
            ("method",    "공법"),
            ("direction", "방향"),
            ("lane",      "차로"),
            ("work_date", "시공일자"),
        ]

        vars_ = {}
        frame = ctk.CTkFrame(dlg, fg_color="transparent")
        frame.pack(padx=20, pady=4, fill="x")
        for field, label in fields:
            row = ctk.CTkFrame(frame, fg_color="transparent")
            row.pack(fill="x", pady=3)
            ctk.CTkLabel(row, text=label, width=160, anchor="w",
                         font=(self.font_family, 12)).pack(side="left")
            var = ctk.StringVar(value=headers[pre_map[field]] if field in pre_map else "(없음)")
            vars_[field] = var
            self._create_styled_option_menu(row, variable=var, values=choices_with_blank,
                                            width=180).pack(side="left", padx=(8, 0))

        result = {}

        def on_ok():
            nonlocal result
            for field, _ in fields:
                chosen = vars_[field].get()
                if chosen != "(없음)" and chosen in headers:
                    result[field] = headers.index(chosen)
            if "start" not in result or "end" not in result:
                self._show_error("필수 항목 누락", "시작 이정과 끝 이정은 반드시 선택해야 합니다.")
                return
            dlg.destroy()

        def on_cancel():
            nonlocal result
            result = None
            dlg.destroy()

        btn_row = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_row.pack(padx=20, pady=(8, 16), fill="x")
        self._create_button(btn_row, text="확인", command=on_ok, width=100).pack(side="left", padx=(0, 8))
        self._create_button(btn_row, text="취소", command=on_cancel, width=100,
                      fg_color="gray30").pack(side="left")

        dlg.wait_window()
        return result if result != {} else None

    def on_export_ps(self):
        # 현재 노선의 캔버스를 대상으로 내보내기
        if self.current_route_index < 0 or self.current_route_index >= len(self.routes):
            self._show_error("노선 없음", "내보낼 노선이 없습니다. 먼저 노선을 추가해 주세요.")
            return
        route = self.routes[self.current_route_index]
        canvas = route.get("canvas")
        if canvas is None:
            self._show_error("오류", "내보낼 캔버스를 찾을 수 없습니다.")
            return
        fpath = filedialog.asksaveasfilename(
            title="모식도 내보내기(PostScript)",
            defaultextension=".ps",
            filetypes=[("PostScript", "*.ps"), ("모든 파일", "*.*")],
            initialfile="schematic.ps",
        )
        if not fpath:
            return
        try:
            canvas.xview_moveto(0.0)
            canvas.update()
            canvas.postscript(file=fpath, colormode="color")
            self._show_info("완료", "모식도를 PostScript로 저장했습니다.\n(PNG 변환은 외부 도구 사용)")
        except Exception as ex:
            self._show_error("오류", f"내보내기 실패: {ex}")

    def on_export_pdf(self):
        """현재 노선의 모식도를 PDF 파일로 내보냅니다."""
        # 0. 필요한 라이브러리 확인
        if not _PDF_LIBS_AVAILABLE:
            self._show_error(
                "라이브러리 필요",
                "PDF 내보내기 기능을 사용하려면 'Pillow'와 'reportlab' 라이브러리가 필요합니다.\n\n"
                "터미널(cmd)을 열고 아래 명령어를 실행하여 설치해 주세요:\n"
                "pip install --upgrade Pillow reportlab\n"
                "pip install Pillow reportlab"
            )
            return

        # 1. 현재 노선 및 캔버스 확인
        if self.current_route_index < 0 or self.current_route_index >= len(self.routes):
            self._show_error("노선 없음", "내보낼 노선이 없습니다.")
            return
        route = self.routes[self.current_route_index]
        canvas = route.get("canvas")
        if canvas is None:
            self._show_error("오류", "내보낼 캔버스를 찾을 수 없습니다.")
            return

        # 2. 저장할 PDF 파일 경로 요청
        fpath = filedialog.asksaveasfilename(
            title="모식도 PDF로 내보내기",
            defaultextension=".pdf",
            filetypes=[("PDF 파일", "*.pdf"), ("모든 파일", "*.*")],
            initialfile=f"{route['name']}_schematic.pdf",
        )
        if not fpath:
            return

        # --- 진행률 팝업 생성 ---
        progress_dlg = self._create_popup_window(self)
        progress_dlg.title("")
        progress_dlg.transient(self)
        progress_dlg.geometry("300x100")
        ctk.CTkLabel(progress_dlg, text="PDF 생성 중...\n잠시만 기다려 주세요.", font=(self.font_family, 14)).pack(pady=20, expand=True)
        progress_dlg.update()

        # --- Ghostscript 번들링 처리 ---
        try:
            from PIL import EpsImagePlugin
            if getattr(sys, 'frozen', False):
                # 번들 내(_MEIPASS) 우선, 없으면 exe 옆 폴더도 확인
                base_dirs = [getattr(sys, '_MEIPASS', ''), os.path.dirname(sys.executable)]
            else:
                # 개발 모드: mixins/의 상위 = 프로젝트 루트
                base_dirs = [os.path.dirname(os.path.dirname(os.path.abspath(__file__)))]
            for base_dir in base_dirs:
                gs_path = os.path.join(base_dir, "ghostscript", "bin", "gswin64c.exe")
                if base_dir and os.path.exists(gs_path):
                    EpsImagePlugin.gs_windows_binary = gs_path
                    break
        except Exception:
            pass

        try:
            # 3. PDF 생성 준비 (A3 가로 방향)
            from reportlab.lib.pagesizes import A3
            pdf = reportlab_canvas.Canvas(fpath, pagesize=landscape(A3))
            page_width, page_height = landscape(A3)
            # 현재 보기 모드 및 상태 플래그 가져오기
            view_mode = self.view_mode.get()
            # PDF에 직접 텍스트를 그리기 위한 좌표 정보를 저장할 리스트
            structure_labels_to_draw = []
            view_hpci = self.view_hpci.get()
            view_di = self.view_di.get()
            view_aar = self.view_aar.get()
            view_rd = self.view_rd.get()
            view_iri = self.view_iri.get()

            # --- 폰트 설정 (문서 시작 시 한 번만) ---
            # 한글 폰트 등록을 시도하고 성공 여부를 플래그로 관리합니다.
            korean_font_name = "MalgunGothic"
            korean_font_available = False
            try:
                from reportlab.pdfbase import pdfmetrics
                from reportlab.pdfbase.ttfonts import TTFont
    
            # 폰트 파일 경로: 번들 내 → 프로젝트 폴더 → Windows 시스템 폰트 순으로 탐색
                font_path = None
                candidates = [
                    os.path.join(get_user_data_dir(), "fonts", "NotoSansKR-Regular.ttf"),
                    resource_path("fonts\\NotoSansKR-Regular.ttf"),
                    resource_path("malgun.ttf"),
                    os.path.join(os.path.dirname(os.path.abspath(__file__)), "malgun.ttf"),
                    r"C:\Windows\Fonts\malgun.ttf",
                ]
                for fp in candidates:
                    if os.path.exists(fp):
                        font_path = fp
                        break
                if font_path is None:
                    raise FileNotFoundError("사용 가능한 한글 폰트를 찾을 수 없습니다.")
                pdfmetrics.registerFont(TTFont(korean_font_name, font_path))
                korean_font_available = True
            except Exception:
                log_exception("PDF 내보내기용 폰트 로드 실패")
                korean_font_available = False

            # 4. 전체 노선을 7km 단위로 분할하여 여러 페이지에 그리기
            r_start = route['start_km']
            r_end = route['end_km']
            segment_km = 7.0  # 한 번에 그릴 길이 (km)

            # 시작점을 1km 단위로 내림하여 구간을 정렬합니다.
            current_km = math.floor(r_start)
            page_num = 1
            while current_km < r_end:
                # --- 한 페이지에 두 개의 세그먼트(위/아래)를 그립니다 ---
                if current_km >= r_end:
                    break

                # 상단 세그먼트 (예: 35.0 ~ 42.0)
                start1 = current_km
                end1 = min(current_km + segment_km, math.ceil(r_end))

                # 하단 세그먼트 (예: 42.0 ~ 49.0)
                start2 = end1
                end2 = start2 + segment_km

                # 그리기 함수는 캔버스를 직접 수정하므로, 임시 캔버스를 사용합니다.
                # 캔버스를 화면 밖에 배치하고 업데이트하여 렌더링을 강제합니다.
                temp_canvas = tk.Canvas(self, width=int(page_width), height=int(page_height/2))
                temp_canvas.place(x=-page_width-200, y=0)

                # PDF 생성용 임시 캔버스에 기본 폰트 설정 (폰트 누락 문제 해결)
                try:
                    font_family = self.font_family
                    try:
                        tk.Label(temp_canvas, font=(font_family, 9)).destroy()
                    except tk.TclError:
                        font_family = "Arial"
                    temp_canvas.option_add("*Font", (font_family, 9))
                except Exception:
                    pass # 폰트 설정 실패 시에도 계속 진행
                temp_canvas.update()

                # 상단 세그먼트 이미지 생성
                self.draw_schematic_for_pdf(temp_canvas, route, start1, end1, page_width, view_mode, r_start, r_end, view_hpci, view_di, view_aar, structure_labels_to_draw, segment_index=0, view_rd=view_rd, view_iri=view_iri)
                temp_canvas.update() # 그리기가 완료되도록 다시 업데이트

                # region을 지정하여 PS 픽셀과 캔버스 픽셀이 정확히 1:scale 매핑되게 한다
                _seg0_cw = next((l['canvas_w'] for l in structure_labels_to_draw if l.get('segment') == 0 and 'canvas_w' in l), int(page_width))
                _seg0_ch = next((l['canvas_h'] for l in structure_labels_to_draw if l.get('segment') == 0 and 'canvas_h' in l), 450)
                ps_data1 = temp_canvas.postscript(colormode="color", x=0, y=0, width=_seg0_cw, height=_seg0_ch)
                img1 = Image.open(io.BytesIO(ps_data1.encode('utf-8')))
                img1.load(scale=4) # 고해상도 렌더링 (오류 발생 시 scale 값을 2 또는 1로 줄여보세요)

                img1_orig_w, img1_orig_h = img1.size  # 크롭 전 원본 이미지 크기 저장
                img1_inverted = ImageOps.invert(img1.convert('RGB'))
                _raw_bbox1 = img1_inverted.getbbox()
                seg0_crop_bbox = (0, 0, img1_orig_w, img1_orig_h)
                if _raw_bbox1:
                    # 구조물명 등 외곽 요소가 잘리는 문제를 해결하기 위해 바운딩 박스에 여백 추가
                    x1, y1, x2, y2 = _raw_bbox1
                    padding = 50
                    seg0_crop_bbox = (max(0, x1 - padding), max(0, y1 - padding),
                                     min(img1_orig_w, x2 + padding), min(img1_orig_h, y2 + padding))
                    img1 = img1.crop(seg0_crop_bbox)

                # PDF에 상단 이미지 그리기
                img_w, img_h = img1.size
                aspect = img_h / float(img_w) if img_w > 0 else 1
                draw_w = page_width * 0.95
                draw_h = draw_w * aspect
                # 상단 영역의 위쪽에 배치하여 아래 세그먼트와의 간격을 확보
                # 페이지의 상단 절반 영역에서 중앙보다 약간 위쪽에 배치
                seg0_pdf_x = page_width * 0.025
                seg0_pdf_y = page_height * 0.5 + 20
                seg0_draw_w = draw_w
                seg0_draw_h = draw_h
                seg0_img_w = img_w
                seg0_img_h = img_h
                pdf.drawImage(ImageReader(img1), seg0_pdf_x, seg0_pdf_y, width=seg0_draw_w, height=seg0_draw_h)

                # 하단 세그먼트 그리기 (end2가 start2보다 클 때만)
                seg1_info_available = False
                seg1_pdf_x = seg1_pdf_y = seg1_draw_w = seg1_draw_h = 0
                seg1_img_w = seg1_img_h = 1
                seg1_crop_bbox = (0, 0, 1, 1)
                if end2 > start2:
                    temp_canvas.delete("all")
                    self.draw_schematic_for_pdf(temp_canvas, route, start2, end2, page_width, view_mode, r_start, r_end, view_hpci, view_di, view_aar, structure_labels_to_draw, segment_index=1, view_rd=view_rd, view_iri=view_iri)
                    temp_canvas.update() # 그리기가 완료되도록 다시 업데이트

                    _seg1_cw = next((l['canvas_w'] for l in structure_labels_to_draw if l.get('segment') == 1 and 'canvas_w' in l), int(page_width))
                    _seg1_ch = next((l['canvas_h'] for l in structure_labels_to_draw if l.get('segment') == 1 and 'canvas_h' in l), 450)
                    ps_data2 = temp_canvas.postscript(colormode="color", x=0, y=0, width=_seg1_cw, height=_seg1_ch)
                    img2 = Image.open(io.BytesIO(ps_data2.encode('utf-8')))
                    img2.load(scale=4) # 고해상도 렌더링
                    img2_orig_w, img2_orig_h = img2.size  # 크롭 전 원본 이미지 크기 저장
                    img2_inverted = ImageOps.invert(img2.convert('RGB'))
                    _raw_bbox2 = img2_inverted.getbbox()
                    seg1_crop_bbox = (0, 0, img2_orig_w, img2_orig_h)
                    if _raw_bbox2:
                        # 구조물명 등 외곽 요소가 잘리는 문제를 해결하기 위해 바운딩 박스에 여백 추가
                        x1, y1, x2, y2 = _raw_bbox2
                        padding = 50
                        seg1_crop_bbox = (max(0, x1 - padding), max(0, y1 - padding),
                                         min(img2_orig_w, x2 + padding), min(img2_orig_h, y2 + padding))
                        img2 = img2.crop(seg1_crop_bbox)

                    # PDF에 하단 이미지 그리기
                    img_w, img_h = img2.size
                    aspect = img_h / float(img_w) if img_w > 0 else 1
                    draw_w = page_width * 0.95
                    draw_h = draw_w * aspect
                    # 하단 영역의 아래쪽에 배치
                    # 페이지의 하단 절반 영역에서 중앙보다 약간 아래쪽에 배치
                    seg1_pdf_x = page_width * 0.025
                    seg1_pdf_y = page_height * 0.5 - draw_h - 20
                    seg1_draw_w = draw_w
                    seg1_draw_h = draw_h
                    seg1_img_w = img_w
                    seg1_img_h = img_h
                    seg1_info_available = True
                    pdf.drawImage(ImageReader(img2), seg1_pdf_x, seg1_pdf_y, width=seg1_draw_w, height=seg1_draw_h)

                # 임시 캔버스 정리
                temp_canvas.destroy()

                # --- 한글 라벨 직접 그리기 (PS/Ghostscript 미지원 한글을 reportlab으로 오버레이) ---
                if korean_font_available and structure_labels_to_draw:

                    # canvas 좌표를 PDF 포인트로 직접 변환 (PS 픽셀 스케일 의존 제거)
                    # Tkinter postscript는 화면 DPI(보통 96)에 따라 72/96=0.75 배율을 적용하므로
                    # PS_SCALE=4 가정이 틀려 텍스트가 우측으로 밀리는 문제를 해결
                    def _canvas_to_pdf(cx, cy, pdf_x, pdf_y, dw, dh, canvas_w, canvas_h):
                        """캔버스 픽셀 좌표 → PDF 포인트 좌표 직접 변환."""
                        rx = max(0.0, min(1.0, cx / canvas_w)) if canvas_w > 0 else 0.5
                        ry = max(0.0, min(1.0, cy / canvas_h)) if canvas_h > 0 else 0.5
                        return pdf_x + rx * dw, pdf_y + (1.0 - ry) * dh

                    for lbl in structure_labels_to_draw:
                        seg_idx = lbl.get('segment', 0)
                        if seg_idx == 0:
                            _px, _py, _dw, _dh = seg0_pdf_x, seg0_pdf_y, seg0_draw_w, seg0_draw_h
                        else:
                            if not seg1_info_available:
                                continue
                            _px, _py, _dw, _dh = seg1_pdf_x, seg1_pdf_y, seg1_draw_w, seg1_draw_h

                        _cw = lbl.get('canvas_w', 1)
                        _ch = lbl.get('canvas_h', 1)

                        color_hex = lbl.get('color', '#000000').lstrip('#')
                        try:
                            cr = int(color_hex[0:2], 16) / 255.0
                            cg = int(color_hex[2:4], 16) / 255.0
                            cb = int(color_hex[4:6], 16) / 255.0
                        except (ValueError, IndexError):
                            cr, cg, cb = 0.0, 0.0, 0.0
                        pdf.setFillColorRGB(cr, cg, cb)

                        lbl_type = lbl.get('type')
                        if lbl_type == 'direction':
                            pdf.setFont(korean_font_name, lbl.get('font_size', 10))
                            px, py = _canvas_to_pdf(lbl['cx'], lbl['cy'], _px, _py, _dw, _dh, _cw, _ch)
                            if lbl.get('anchor') == 'w':
                                pdf.drawString(px, py, lbl['text'])
                            else:
                                pdf.drawCentredString(px, py, lbl['text'])
                        elif lbl_type == 'ic_name':
                            pdf.setFont(korean_font_name, lbl.get('font_size', 11))
                            char_h = lbl.get('char_h', 16)
                            for i, char in enumerate(lbl['text']):
                                cy_char = lbl['cy_start'] + i * char_h
                                px, py = _canvas_to_pdf(lbl['cx'], cy_char, _px, _py, _dw, _dh, _cw, _ch)
                                pdf.drawCentredString(px, py, char)
                        elif lbl_type == 'structure':
                            fs = lbl.get('font_size', 8)
                            pdf.setFont(korean_font_name, fs)
                            name_str = lbl.get('text', '')
                            cx_c = lbl['cx']
                            cy_c = lbl['cy']
                            line_h = fs * 1.4
                            total_text_h = len(name_str) * line_h
                            start_cy = cy_c - total_text_h / 2 + line_h / 2
                            for i, char in enumerate(name_str):
                                cy_char = start_cy + i * line_h
                                px, py = _canvas_to_pdf(cx_c, cy_char, _px, _py, _dw, _dh, _cw, _ch)
                                pdf.drawCentredString(px, py, char)

                # --- 헤더 및 푸터 추가 ---
                # 폰트 등록 성공 여부에 따라 폰트를 설정합니다.
                if korean_font_available:
                    pdf.setFont(korean_font_name, 20)
                else:
                    pdf.setFont("Helvetica", 20)

                pdf.setFillColorRGB(0, 0, 0) # 헤더를 그리기 전에 검은색으로 설정
                # 헤더: 노선명 및 구간 정보
                header_text = f"노선: {route['name']}  |  구간: {route['start_km']:.2f}km ~ {route['end_km']:.2f}km"
                pdf.drawCentredString(page_width / 2, page_height - 40, header_text)

                directions_text = f"({route['directions'][0]} / {route['directions'][1]})"
                pdf.drawCentredString(page_width / 2, page_height - 65, directions_text)
                # 푸터: 페이지 번호
                pdf.setFont("Helvetica", 12) # 페이지 번호 폰트 크기 조정
                pdf.drawCentredString(page_width / 2, 30, f"- {page_num} -")

                pdf.showPage() # 다음 페이지로
                page_num += 1
                current_km = end2 # 다음 루프를 위해 시작 km 업데이트
                structure_labels_to_draw.clear() # 다음 페이지를 위해 리스트 초기화
                progress_dlg.update() # 페이지 완료 시마다 UI 응답 유지

            pdf.save()

        except PermissionError:
            progress_dlg.destroy()
            self._show_error(
                "권한 오류 (Permission Denied)",
                f"PDF 파일 저장에 실패했습니다.\n\n파일이 다른 프로그램에서 열려 있거나, 해당 위치에 저장할 권한이 없는지 확인해 주세요.\n\n파일 경로: {fpath}"
            )
        except Exception as ex:
            progress_dlg.destroy()
            self._show_error("오류", f"PDF 내보내기 실패: {ex}\n\nGhostscript가 설치되어 있는지 확인해 주세요.")
        else:
            progress_dlg.destroy()
            self._show_info("완료", "모식도를 PDF 파일로 저장했습니다.")

    def on_export_all_to_excel(self):
        """모든 노선의 본선 및 램프 이력을 하나의 Excel 파일로 내보냅니다."""
        if not _EXCEL_LIBS_AVAILABLE:
            self._show_error(
                "라이브러리 필요",
                "Excel 내보내기 기능을 사용하려면 'openpyxl' 라이브러리가 필요합니다.\n\n"
                "터미널(cmd)을 열고 아래 명령어를 실행하여 설치해 주세요:\n"
                "pip install openpyxl"
            )
            return

        fpath = filedialog.asksaveasfilename(
            title="모든 이력 Excel로 내보내기",
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")],
            initialfile="전체노선_이력_데이터.xlsx",
        )
        if not fpath:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "전체 이력"

            headers = [
                "노선명", "구분", "IC/JCT명", "램프명", "방향",
                "시작이정(km)", "종료이정(km)", "연장(m)", "차로", "공법", "시공일자"
            ]
            ws.append(headers)

            # 헤더 스타일
            header_font = Font(bold=True)
            center_align = Alignment(horizontal='center', vertical='center')
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

            # 데이터 수집 및 추가
            all_data = []

            # 1. 본선 이력 수집
            for route in self.routes:
                for entry in route.get("entries", []):
                    length_m = (entry.get('end', 0.0) - entry.get('start', 0.0)) * 1000
                    all_data.append([
                        route.get('name', ''), '본선', '', '', entry.get('direction', ''),
                        entry.get('start', 0.0), entry.get('end', 0.0), f"{length_m:.1f}",
                        entry.get('lane', ''), entry.get('method', ''), entry.get('work_date', '')
                    ])

            # 2. 램프 이력 수집
            for ic in getattr(self, 'ics', []):
                for ramp in ic.get('ramps', []):
                    for entry in ramp.get('entries', []):
                        length_m = (entry.get('end', 0.0) - entry.get('start', 0.0)) * 1000
                        all_data.append([
                            ic.get('route', ''), '램프', ic.get('name', ''), ramp.get('name', ''), '',
                            entry.get('start', 0.0), entry.get('end', 0.0), f"{length_m:.1f}",
                            entry.get('lane', ''), entry.get('method', ''), entry.get('work_date', '')
                        ])

            # 노선명, 시공일자 순으로 정렬
            all_data.sort(key=lambda x: (x[0], str(x[10])))

            # 워크시트에 데이터 쓰기
            for row_data in all_data:
                ws.append(row_data)

            # 열 너비 자동 조정
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(fpath)
            self._show_info("성공", f"모든 이력 데이터를 Excel 파일로 저장했습니다:\n{fpath}")

        except Exception as ex:
            self._show_error("오류", f"Excel 파일 저장 중 오류가 발생했습니다:\n{ex}")

    def on_menu_save_route(self):
        # 모든 노선 자동 저장 (파일 대화상자 없이 동일 폴더에 저장)
        # CSV 2종 저장: (1) 이력, (2) IC 정보
        base_dir = get_user_data_dir()

        saved_files = []
        year_to_condition_rows = {} # 연도별 포장상태 데이터 수집용
        
        try:
            # 1. 통합 이력 저장 (all_maintenance_history.csv)
            hist_fname = "all_maintenance_history.csv"
            hist_fpath = os.path.join(base_dir, hist_fname)
            with open(hist_fpath, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                # 'year' 컬럼 추가
                writer.writerow([
                    "route_name", "route_start_km", "route_end_km",
                    "start_km", "end_km", "method", "direction", "lane", "work_date", "year", "timestamp",
                    "row_type", "route_dir1", "route_dir2"
                ])
                
                for route in self.routes:
                    dirs = route.get("directions") or DIRECTIONS
                    dir1 = dirs[0]
                    dir2 = dirs[1 if len(dirs) > 1 else 0]
                    entries = route.get("entries", [])
                    
                    if not entries:
                        # 메타 데이터만 있는 행
                        writer.writerow([
                            route.get("name", ""), route.get("start_km", 0.0), route.get("end_km", 0.0),
                            "", "", "", "", "", "", "", "",
                            "meta", dir1, dir2
                        ])
                    else:
                        for it in entries:
                            if it.get("plan"):
                                continue  # 사업계획 구간은 all_business_plan.csv에 별도 저장
                            wd = str(it.get("work_date") or "")
                            year = wd[:4] if len(wd) >= 4 else ""
                            writer.writerow([
                                route.get("name", ""), route.get("start_km", 0.0), route.get("end_km", 0.0),
                                it.get("start", 0.0), it.get("end", 0.0), it.get("method", ""),
                                it.get("direction", ""), it.get("lane", ""), wd, year, it.get("ts", ""),
                                "entry", dir1, dir2
                            ])
            saved_files.append(hist_fname)

            # 2. 통합 IC 저장 (all_ic.csv)
            ic_fname = "all_ic.csv"
            ic_fpath = os.path.join(base_dir, ic_fname)
            with open(ic_fpath, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(["route_name", "ic_name", "ic_km", "ramp_name", "ramp_km"])

                for ic in getattr(self, "ics", []):
                    rname = ic.get("route", "")
                    name = ic.get("name", "")
                    km = float(ic.get("km", 0.0))
                    ramps = ic.get("ramps") or []
                    if not ramps:
                        writer.writerow([rname, name, km, "", ""])
                    else:
                        for r in ramps:
                            if isinstance(r, dict):
                                nm = r.get("name", "")
                                try:
                                    rkm = float(r.get("km", km))
                                except Exception:
                                    rkm = km
                                writer.writerow([rname, name, km, nm, rkm])
            saved_files.append(ic_fname)

            # 4. 통합 구조물 저장 (all_structures.csv)
            struct_fname = "all_structures.csv"
            struct_fpath = os.path.join(base_dir, struct_fname)
            with open(struct_fpath, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(["route_name", "name", "type", "direction", "start_km", "length_m"])
                for route in self.routes:
                    for s in route.get("structures", []):
                        length = s.get('length', 0.0)
                        if length == 0.0 and s.get('end') and s.get('start'):
                            length = (s['end'] - s['start']) * 1000.0
                        writer.writerow([
                            route.get("name", ""), s.get("name", ""), s.get("type", ""),
                            s.get("direction", "양방향"), s.get("start", 0.0), length
                        ])
            saved_files.append(struct_fname)

            # 5. 통합 AAR 데이터 저장 (all_aar_data.csv)
            aar_fname = "all_aar_data.csv"
            aar_fpath = os.path.join(base_dir, aar_fname)
            with open(aar_fpath, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(["route_name", "direction", "lane", "start_km", "year", "aar_grade"])
                for route in self.routes:
                    aar_data = route.get("aar_data", {})
                    for (d, l, skm), val_map in aar_data.items():
                        if isinstance(val_map, dict):
                            for y, val in val_map.items():
                                writer.writerow([route.get("name", ""), d, l, skm, y, val])
                        else:
                            writer.writerow([route.get("name", ""), d, l, skm, datetime.now().strftime("%Y"), val_map])
            saved_files.append(aar_fname)

            # 6. 통합 지사 이력 저장 (all_branch_history.csv)
            branch_hist_fname = "all_branch_history.csv"
            branch_hist_fpath = os.path.join(base_dir, branch_hist_fname)
            with open(branch_hist_fpath, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow([
                    "route_name", "route_start_km", "route_end_km",
                    "start_km", "end_km", "method", "direction", "lane", "work_date", "timestamp"
                ])
                for route in self.routes:
                    for it in route.get("branch_entries", []):
                        writer.writerow([
                            route.get("name", ""), route.get("start_km", 0.0), route.get("end_km", 0.0),
                            it.get("start", 0.0), it.get("end", 0.0), it.get("method", ""),
                            it.get("direction", ""), it.get("lane", ""), it.get("work_date", ""), it.get("ts", "")
                        ])
            saved_files.append(branch_hist_fname)

            # 7. 포장상태불량지수 데이터 수집 (연도별 통합 저장 - 기존 로직 유지하되 전체 노선 순회)
            for route in self.routes:
                di_data = route.get('di_data', {})
                hpci_data = route.get('hpci_data', {})
                rd_data = route.get('rd_data', {})
                iri_data = route.get('iri_data', {})
                pavement_data = route.get('pavement_data', {})

                all_keys = set()
                for data_map in [di_data, hpci_data, rd_data, iri_data]:
                    for k, v_map in data_map.items():
                        if isinstance(v_map, dict):
                            all_keys.add(k)
                for k in pavement_data:
                    all_keys.add(k)

                for (d, l, skm_str) in all_keys:
                    years_here = set()
                    for data_map in [di_data, hpci_data, rd_data, iri_data]:
                        val_map = data_map.get((d, l, skm_str), {})
                        if isinstance(val_map, dict):
                            years_here.update(val_map.keys())
                    if not years_here:
                        years_here = {""}

                    pav_val = pavement_data.get((d, l, skm_str), "")

                    for year in years_here:
                        hpci_val = hpci_data.get((d, l, skm_str), {}).get(year, "")
                        di_val = di_data.get((d, l, skm_str), {}).get(year, "")
                        rd_val = rd_data.get((d, l, skm_str), {}).get(year, "")
                        iri_val = iri_data.get((d, l, skm_str), {}).get(year, "")

                        try:
                            skm = float(skm_str)
                            ekm = skm + GRID_KM
                        except Exception:
                            skm, ekm = 0.0, 0.0

                        if year not in year_to_condition_rows:
                            year_to_condition_rows[year] = []

                        year_to_condition_rows[year].append([
                            route.get('name', ''), d, l, skm, f"{ekm:.2f}", hpci_val, di_val, rd_val, iri_val, pav_val
                        ])

            # 모든 노선 순회 후 포장상태불량지수 단일 파일 저장 (all_condition.csv, year 컬럼 포함)
            cond_fname = "all_condition.csv"
            cond_fpath = os.path.join(base_dir, cond_fname)
            with open(cond_fpath, "w", newline="", encoding="utf-8-sig") as f_cond:
                writer = csv.writer(f_cond)
                writer.writerow(["route_name", "direction", "lane", "start_km", "end_km", "year", "hpci", "di", "rd", "iri", "포장형식"])
                # year_to_condition_rows[year] = [[route_name, d, l, skm, ekm, hpci, di, rd, iri, pav], ...]
                all_cond_rows = []
                for year, rows in year_to_condition_rows.items():
                    for row in rows:
                        # row = [route_name, d, l, skm, ekm, hpci, di, rd, iri, pav]
                        all_cond_rows.append([row[0], row[1], row[2], row[3], row[4],
                                              year, row[5], row[6], row[7], row[8], row[9] if len(row) > 9 else ""])
                all_cond_rows.sort(key=lambda x: (x[0], x[1], x[2], float(x[3]), x[5]))
                for row in all_cond_rows:
                    writer.writerow(row)
            saved_files.append(cond_fname)

        # 공법(METHOD_STYLES) 저장 (모든 노선에 대해 한 번만)
            try:
                method_fname = "all_routes_methods.csv"
                method_fpath = os.path.join(base_dir, method_fname)
                with open(method_fpath, 'w', newline='', encoding='utf-8-sig') as fm:
                    mwriter = csv.writer(fm)
                    mwriter.writerow(['method_name', 'fill_color', 'category'])
                    for name, style in METHOD_STYLES.items():
                        mwriter.writerow([name, style.get('fill', '#718096'),
                                          METHOD_CATEGORY_MAP.get(name, '')])
                if method_fname not in saved_files:
                    saved_files.append(method_fname)
            except Exception:
                pass

            # 하자기간 설정 저장
            try:
                w_fname = "warranty_settings.csv"
                w_fpath = os.path.join(base_dir, w_fname)
                with open(w_fpath, "w", newline="", encoding="utf-8-sig") as fw:
                    ww = csv.writer(fw)
                    ww.writerow(["category", "period", "rate"])
                    for cat, info in CATEGORY_WARRANTY.items():
                        ww.writerow([cat, info.get("period", 3), info.get("rate", 100.0)])
                if w_fname not in saved_files:
                    saved_files.append(w_fname)
            except Exception:
                pass

            try:
                self._write_sqlite_cache(base_dir)
                saved_files.append("highway_data.db")
            except Exception:
                log_exception("SQLite 캐시 저장 실패")
        except Exception as ex:
            self._show_error("오류", f"저장 실패: {ex}")

    def _write_sqlite_cache(self, base_dir: str):
        db_path = os.path.join(base_dir, "highway_data.db")
        source_signature = build_source_signature(base_dir)
        conn = sqlite3.connect(db_path)
        try:
            cur = conn.cursor()
            cur.executescript("""
            PRAGMA journal_mode=WAL;
            PRAGMA synchronous=NORMAL;
            DROP TABLE IF EXISTS metadata;
            DROP TABLE IF EXISTS routes;
            DROP TABLE IF EXISTS entries;
            DROP TABLE IF EXISTS ics;
            DROP TABLE IF EXISTS structures;
            DROP TABLE IF EXISTS aar;
            DROP TABLE IF EXISTS conditions;
            DROP TABLE IF EXISTS methods;
            DROP TABLE IF EXISTS warranty;

            CREATE TABLE metadata (
                key TEXT PRIMARY KEY,
                value TEXT
            );
            CREATE TABLE routes (
                route_name TEXT PRIMARY KEY,
                start_km REAL,
                end_km REAL,
                dir1 TEXT,
                dir2 TEXT,
                lane_count INTEGER
            );
            CREATE TABLE entries (
                route_name TEXT,
                entry_type TEXT,
                start_km REAL,
                end_km REAL,
                method TEXT,
                direction TEXT,
                lane TEXT,
                work_date TEXT,
                ts TEXT
            );
            CREATE TABLE ics (
                route_name TEXT,
                ic_name TEXT,
                ic_km REAL,
                ramp_name TEXT,
                ramp_km REAL
            );
            CREATE TABLE structures (
                route_name TEXT,
                name TEXT,
                type TEXT,
                direction TEXT,
                start_km REAL,
                end_km REAL,
                length_m REAL
            );
            CREATE TABLE aar (
                route_name TEXT,
                direction TEXT,
                lane TEXT,
                start_km REAL,
                year TEXT,
                aar_grade INTEGER
            );
            CREATE TABLE conditions (
                route_name TEXT,
                direction TEXT,
                lane TEXT,
                start_km REAL,
                year TEXT,
                hpci INTEGER,
                di REAL,
                rd REAL,
                iri REAL,
                pavement TEXT
            );
            CREATE TABLE methods (
                method_name TEXT PRIMARY KEY,
                fill_color TEXT,
                category TEXT
            );
            CREATE TABLE warranty (
                category TEXT PRIMARY KEY,
                period INTEGER,
                rate REAL
            );
            CREATE INDEX idx_entries_route_type ON entries(route_name, entry_type);
            CREATE INDEX idx_conditions_route_year ON conditions(route_name, year);
            CREATE INDEX idx_aar_route_year ON aar(route_name, year);
            """)

            route_rows = []
            entry_rows = []
            ic_rows = []
            struct_rows = []
            aar_rows = []
            condition_rows = []

            for route in self.routes:
                dirs = route.get("directions") or DIRECTIONS
                dir1 = dirs[0] if dirs else ""
                dir2 = dirs[1] if len(dirs) > 1 else dir1
                route_rows.append((
                    route.get("name", ""),
                    float(route.get("start_km", 0.0)),
                    float(route.get("end_km", 0.0)),
                    dir1, dir2,
                    int(route.get("lane_count", 4)),
                ))

                for it in route.get("entries", []):
                    if it.get("plan"):
                        continue  # 사업계획 구간은 캐시 대신 all_business_plan.csv에 저장
                    entry_rows.append((
                        route.get("name", ""), "main",
                        float(it.get("start", 0.0)), float(it.get("end", 0.0)),
                        str(it.get("method", "")), str(it.get("direction", "")),
                        str(it.get("lane", "")), str(it.get("work_date", "")),
                        str(it.get("ts", "")),
                    ))
                for it in route.get("branch_entries", []):
                    entry_rows.append((
                        route.get("name", ""), "branch",
                        float(it.get("start", 0.0)), float(it.get("end", 0.0)),
                        str(it.get("method", "")), str(it.get("direction", "")),
                        str(it.get("lane", "")), str(it.get("work_date", "")),
                        str(it.get("ts", "")),
                    ))

                for s in route.get("structures", []):
                    start_km = float(s.get("start", 0.0))
                    end_km = float(s.get("end", start_km))
                    length_m = float(s.get("length", (end_km - start_km) * 1000.0))
                    struct_rows.append((
                        route.get("name", ""),
                        str(s.get("name", "")),
                        str(s.get("type", "")),
                        str(s.get("direction", "양방향")),
                        start_km, end_km, length_m
                    ))

                for (d, l, skm), val_map in route.get("aar_data", {}).items():
                    if isinstance(val_map, dict):
                        for y, val in val_map.items():
                            aar_rows.append((route.get("name", ""), d, l, float(skm), str(y), int(val)))

                di_data = route.get("di_data", {})
                hpci_data = route.get("hpci_data", {})
                rd_data = route.get("rd_data", {})
                iri_data = route.get("iri_data", {})
                pavement_data = route.get("pavement_data", {})
                all_keys = set(di_data.keys()) | set(hpci_data.keys()) | set(rd_data.keys()) | set(iri_data.keys()) | set(pavement_data.keys())
                for (d, l, skm) in all_keys:
                    years = set()
                    for dmap in (di_data, hpci_data, rd_data, iri_data):
                        vmap = dmap.get((d, l, skm), {})
                        if isinstance(vmap, dict):
                            years.update(vmap.keys())
                    if not years:
                        years = {""}
                    pav = pavement_data.get((d, l, skm), "")
                    for y in years:
                        condition_rows.append((
                            route.get("name", ""), d, l, float(skm), str(y),
                            hpci_data.get((d, l, skm), {}).get(y, None),
                            di_data.get((d, l, skm), {}).get(y, None),
                            rd_data.get((d, l, skm), {}).get(y, None),
                            iri_data.get((d, l, skm), {}).get(y, None),
                            pav
                        ))

            for ic in getattr(self, "ics", []):
                rname = str(ic.get("route", ""))
                iname = str(ic.get("name", ""))
                ikm = float(ic.get("km", 0.0))
                ramps = ic.get("ramps") or []
                if not ramps:
                    ic_rows.append((rname, iname, ikm, "", None))
                else:
                    for r in ramps:
                        if isinstance(r, dict):
                            ic_rows.append((rname, iname, ikm, str(r.get("name", "")), float(r.get("km", ikm))))

            cur.executemany("INSERT INTO routes VALUES (?, ?, ?, ?, ?, ?)", route_rows)
            cur.executemany("INSERT INTO entries VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)", entry_rows)
            cur.executemany("INSERT INTO ics VALUES (?, ?, ?, ?, ?)", ic_rows)
            cur.executemany("INSERT INTO structures VALUES (?, ?, ?, ?, ?, ?, ?)", struct_rows)
            cur.executemany("INSERT INTO aar VALUES (?, ?, ?, ?, ?, ?)", aar_rows)
            cur.executemany("INSERT INTO conditions VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", condition_rows)
            cur.executemany(
                "INSERT INTO methods VALUES (?, ?, ?)",
                [(name, style.get("fill", "#718096"), METHOD_CATEGORY_MAP.get(name, "")) for name, style in METHOD_STYLES.items()]
            )
            cur.executemany(
                "INSERT INTO warranty VALUES (?, ?, ?)",
                [(cat, info.get("period", 3), info.get("rate", 100.0)) for cat, info in CATEGORY_WARRANTY.items()]
            )
            cur.executemany(
                "INSERT INTO metadata VALUES (?, ?)",
                [
                    ("schema_version", str(CACHE_SCHEMA_VERSION)),
                    ("source_signature", source_signature),
                    ("saved_at", datetime.now().isoformat(timespec="seconds")),
                ]
            )
            conn.commit()
        finally:
            conn.close()

    def _load_from_sqlite_cache(self, base_dir: str) -> bool:
        db_path = os.path.join(base_dir, "highway_data.db")
        if not os.path.exists(db_path):
            return False
        conn = sqlite3.connect(db_path)
        try:
            cur = conn.cursor()
            tables = {row[0] for row in cur.execute("SELECT name FROM sqlite_master WHERE type='table'")}
            if "routes" not in tables or "metadata" not in tables:
                return False
            metadata = dict(cur.execute("SELECT key, value FROM metadata"))
            if metadata.get("schema_version") != str(CACHE_SCHEMA_VERSION):
                return False
            if metadata.get("source_signature") != build_source_signature(base_dir):
                return False

            self.routes.clear()
            self.ics = []
            self.current_route_index = -1
            self.canvas_to_route_index = {}

            reset_method_settings()

            if "methods" in tables:
                for name, color, category in cur.execute("SELECT method_name, fill_color, category FROM methods"):
                    METHOD_STYLES[str(name)] = {"fill": str(color or "#718096")}
                    if category:
                        METHOD_CATEGORY_MAP[str(name)] = str(category)

            if "warranty" in tables:
                for cat, period, rate in cur.execute("SELECT category, period, rate FROM warranty"):
                    CATEGORY_WARRANTY[str(cat)] = {"period": int(period or 3), "rate": float(rate or 100.0)}

            route_map = {}
            for route_name, start_km, end_km, dir1, dir2, lane_count in cur.execute(
                "SELECT route_name, start_km, end_km, dir1, dir2, lane_count FROM routes ORDER BY route_name"
            ):
                dirs = [str(dir1 or DIRECTIONS[0]), str(dir2 or DIRECTIONS[1])]
                self.add_route(str(route_name), float(start_km or 0.0), float(end_km or 0.0), dirs)
                route = self.routes[self.current_route_index]
                route["lane_count"] = int(lane_count or 4)
                route_map[str(route_name)] = route

            if "entries" in tables:
                for route_name, entry_type, start_km, end_km, method, direction, lane, work_date, ts in cur.execute(
                    "SELECT route_name, entry_type, start_km, end_km, method, direction, lane, work_date, ts FROM entries ORDER BY route_name, start_km, end_km"
                ):
                    route = route_map.get(str(route_name))
                    if route is None:
                        continue
                    target_list = route["branch_entries"] if str(entry_type) == "branch" else route["entries"]
                    target_list.append({
                        "start": float(start_km or 0.0),
                        "end": float(end_km or 0.0),
                        "method": str(method or ""),
                        "direction": str(direction or ""),
                        "lane": str(lane or "전차로"),
                        "work_date": str(work_date or ""),
                        "ts": str(ts or ""),
                    })

            if "structures" in tables:
                for route_name, name, stype, direction, start_km, end_km, length_m in cur.execute(
                    "SELECT route_name, name, type, direction, start_km, end_km, length_m FROM structures ORDER BY route_name, start_km"
                ):
                    route = route_map.get(str(route_name))
                    if route is None:
                        continue
                    route["structures"].append({
                        "name": str(name or ""),
                        "type": str(stype or ""),
                        "direction": str(direction or "양방향"),
                        "start": float(start_km or 0.0),
                        "end": float(end_km or 0.0),
                        "length": float(length_m or 0.0),
                    })

            if "aar" in tables:
                for route_name, direction, lane, start_km, year, grade in cur.execute(
                    "SELECT route_name, direction, lane, start_km, year, aar_grade FROM aar"
                ):
                    route = route_map.get(str(route_name))
                    if route is None:
                        continue
                    key = (str(direction or ""), str(lane or ""), f"{float(start_km or 0.0):.2f}")
                    route["aar_data"].setdefault(key, {})[str(year or "")] = int(grade or 0)

            if "conditions" in tables:
                for route_name, direction, lane, start_km, year, hpci, di, rd, iri, pavement in cur.execute(
                    "SELECT route_name, direction, lane, start_km, year, hpci, di, rd, iri, pavement FROM conditions"
                ):
                    route = route_map.get(str(route_name))
                    if route is None:
                        continue
                    key = (str(direction or ""), str(lane or ""), f"{float(start_km or 0.0):.2f}")
                    year = str(year or "")
                    if hpci is not None:
                        route["hpci_data"].setdefault(key, {})[year] = int(hpci)
                    if di is not None:
                        route["di_data"].setdefault(key, {})[year] = float(di)
                    if rd is not None:
                        route["rd_data"].setdefault(key, {})[year] = float(rd)
                    if iri is not None:
                        route["iri_data"].setdefault(key, {})[year] = float(iri)
                    if pavement:
                        route.setdefault("pavement_data", {})[key] = str(pavement)

            if "ics" in tables:
                grouped = {}
                for route_name, ic_name, ic_km, ramp_name, ramp_km in cur.execute(
                    "SELECT route_name, ic_name, ic_km, ramp_name, ramp_km FROM ics ORDER BY route_name, ic_km, ic_name"
                ):
                    key = (str(route_name or ""), str(ic_name or ""), float(ic_km or 0.0))
                    grp = grouped.setdefault(key, {
                        "route": str(route_name or ""),
                        "name": str(ic_name or ""),
                        "km": float(ic_km or 0.0),
                        "ramps": []
                    })
                    if ramp_name:
                        grp["ramps"].append({"name": str(ramp_name), "km": float(ramp_km or ic_km or 0.0)})
                self.ics = list(grouped.values())

            if self.routes:
                self.current_route_index = 0
                self.notebook.set(self.routes[0]["name"])
                self.refresh_route_controls_from_route(self.routes[0])

            for route in self.routes:
                route["_data_version"] = int(route.get("_data_version", 0)) + 1
                route["_year_values_cache"] = None
                route["_year_values_version"] = -1
            self._mark_route_data_changed()
            self.update_method_combobox_values()
            self.render_legend()
            self.update_year_filter_values()
            self.draw_schematic()
            return True
        finally:
            conn.close()

