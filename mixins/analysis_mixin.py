# -*- coding: utf-8 -*-
"""고속도로 포장유지보수 이력 관리 - 분석: DI/HPCI/AAR/RD/IRI, 우선순위, 결함리스크, 이동
"""
import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import csv
import customtkinter as ctk
import math
import sqlite3
import json
import logging
from datetime import datetime
import tkinter.font as tkfont
import io
import ctypes
import tempfile

try:
    from PIL import Image, ImageOps, ImageTk
    from reportlab.pdfgen import canvas as reportlab_canvas
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.utils import ImageReader
    _PDF_LIBS_AVAILABLE = True
except ImportError:
    _PDF_LIBS_AVAILABLE = False
    ImageTk = None

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    _EXCEL_LIBS_AVAILABLE = True
except ImportError:
    _EXCEL_LIBS_AVAILABLE = False

try:
    from CTkMessagebox import CTkMessagebox
except ImportError:
    CTkMessagebox = None

# 개별 파일을 직접 실행하거나 작업 폴더가 달라도 프로젝트 루트(상위 폴더)의
# constants/utils 등을 찾을 수 있도록 sys.path 에 추가
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from constants import *
from utils import (
    get_logger, log_exception, log_warning,
    read_csv_with_encoding, get_user_data_dir, resource_path,
    load_method_settings, reset_method_settings,
    get_method_warranty_period, build_source_signature,
    km_floor_to_grid, km_ceil_to_grid, clamp, fmt_km,
    intervals_overlap, lane_conflict,
)
from canvas_utils import (
    _split_gaps, _merge_gaps, _draw_hatch_rect, _compute_overlap_status,
)


class AnalysisMixin:
    def on_manage_di(self):
        self._open_condition_data_dialog('di_data', 'DI 지수', value_type=float,
                                         redraw_on_save=True, csv_enabled=True)

    def on_manage_hpci(self):
        self._open_condition_data_dialog('hpci_data', 'HPCI 등급', value_type=int, redraw_on_save=True)

    def on_manage_aar(self):
        self._open_condition_data_dialog('aar_data', 'AAR 등급', value_type=int, redraw_on_save=True)

    def on_manage_rd(self):
        self._open_condition_data_dialog('rd_data', 'RD 등급', value_type=float, redraw_on_save=False)

    def on_manage_iri(self):
        self._open_condition_data_dialog('iri_data', 'IRI 등급', value_type=float, redraw_on_save=False)

    def on_improvement_priority(self, apply_label=None, apply_callback=None,
                                exclude_sections=None, default_year=None):
        """개량 우선순위 선정 (고도화: 구조물/이력 실좌표 정밀 절단 Trimming + 방향 매칭 개선)

        매개변수(선택):
            apply_label    : 지정하면 '선택 적용' 버튼을 추가하고 그 텍스트로 사용
            apply_callback : 선택 적용 시 호출. callback(items:list, plan_year:str)
            exclude_sections: 후보에서 제외할 구간 목록. 각 항목 dict {route,direction,lane,start,end}
            default_year   : 계획 연도 입력칸 기본값(문자열)
        """
        # 다른 컨텍스트(사업계획/운영계획변경)에서 재호출될 수 있으므로
        # 기존 창이 열려 있으면 닫고 새로 구성한다.
        if hasattr(self, 'priority_window') and self.priority_window is not None and self.priority_window.winfo_exists():
            if apply_callback is None:
                self.priority_window.lift()
                self.priority_window.focus_force()
                return
            try:
                self.priority_window.destroy()
            except Exception:
                pass

        # 1. 윈도우 설정
        dlg = self._create_popup_window(self)
        self.priority_window = dlg
        dlg.title("")
        dlg.geometry("1650x650")
        dlg.transient(self)
        dlg.lift()
        dlg.focus_force()
        
        # 2. 상단 컨트롤 패널
        top_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        top_frame.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(top_frame, text="계획 연도:", font=(self.font_family, 14, "bold")).pack(side="left")
        
        current_yr = int(datetime.now().strftime("%Y"))
        _def_year = str(default_year) if default_year else str(current_yr + 1)
        var_plan_year = tk.StringVar(value=_def_year)
        ent_year = ctk.CTkEntry(top_frame, textvariable=var_plan_year, width=60, font=(self.font_family, 12))
        ent_year.pack(side="left", padx=5)
        
        GAP_TOLERANCE = 0.3

        def show_algorithm_info():
            try:
                plan_y = int(var_plan_year.get())
                overlay_range  = f"{plan_y - 3}~{plan_y - 1}년"
                surface_range  = f"{plan_y - 2}~{plan_y - 1}년"
                score_range    = f"{plan_y - 3}~{plan_y - 1}년"
            except Exception:
                plan_y = 0
                overlay_range  = "계획연도 직전 3년"
                surface_range  = "계획연도 직전 2년"
                score_range    = "계획연도 직전 3년"

            info_text = (
                f"■ 개량 대상 선정 알고리즘 (Plan: {var_plan_year.get()}년)\n\n"
                f"1. [방해물 수집]\n"
                f"   - 구조물(교량 등) 전체 제외\n"
                f"   - 덧씌우기 계열: {overlay_range} 시공 구간 제외 (하자보수 3년)\n"
                f"   - 표면개량 계열: {surface_range} 시공 구간 제외 (하자보수 2년)\n\n"
                f"2. [최대 연장 구간 분리]\n"
                f"   - 방해물을 기준으로 clean 구간 분리 (연장 최대화)\n"
                f"   - 각 clean 구간의 전년도({plan_y - 1}년) 평균 DI ≥ 5.0 확인\n"
                f"   - 평균 DI < 5.0이면 해당 구간 제외\n"
                f"   - 최소 300m 이상인 구간만 개량대상 선정\n\n"
                f"3. [우선순위 산정]\n"
                f"   - AAR등급 평균 ≥ 2인 구간: 별도 관리 (맨 마지막, 초록색)\n"
                f"   - 일반 구간: ① 3개년({score_range}) DI평균 ↓  ② HPCI 3개년 평균 ↓  ③ 연장 ↓\n"
                f"   - 적용공법: 절삭 덧씌우기\n"
            )
            info_dlg = self._create_popup_window(dlg)
            info_dlg.title("")
            info_dlg.geometry("900x380")
            info_dlg.transient(dlg)
            info_dlg.grab_set()
            info_dlg.lift()
            info_dlg.focus_force()
            txt = ctk.CTkTextbox(info_dlg, font=(self.font_family, 12), wrap="word")
            txt.pack(fill="both", expand=True, padx=16, pady=(14, 8))
            txt.insert("end", info_text)
            txt.configure(state="disabled")
            self._create_close_button(info_dlg, info_dlg.destroy, width=100, height=32).pack(pady=(0, 12))

        def on_export_priority_excel():
            if not _EXCEL_LIBS_AVAILABLE:
                self._show_error("라이브러리 필요", "openpyxl 필요")
                return
            if not last_calculated_data:
                self._show_info("알림", "데이터 없음")
                return
            
            fpath = filedialog.asksaveasfilename(
                title="엑셀 저장", defaultextension=".xlsx",
                filetypes=[("Excel 파일", "*.xlsx")], 
                initialfile=f"{var_plan_year.get()}년_개량우선순위_정밀.xlsx", 
                parent=dlg
            )
            if not fpath: return

            try:
                from openpyxl.styles import PatternFill
                wb = Workbook()
                ws = wb.active; ws.title = "우선순위"
                
                headers = ["순위", "노선명", "방향", "차로", "시작(km)", "끝(km)", "연장(m)", "포장형식", "적용공법", "3개년DI", "작년DI", "AAR", "IRI", "RD", "사유"]
                ws.append(headers)

                header_font = Font(bold=True, color="FFFFFF")
                fill_style = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                for cell in ws[1]:
                    cell.font = header_font; cell.fill = fill_style; cell.alignment = Alignment(horizontal='center')

                def _fmt_val(v):
                    if v is None: return ""
                    try: return float(f"{v:.2f}")
                    except Exception: return v

                for item in last_calculated_data:
                    row_vals = {c: get_formatted_val(item, c) for c in cols}
                    if any(row_vals[c] not in active_filters[c] for c in active_filters): continue
                    ws.append([
                        item['rank'], item['route'],
                        get_formatted_val(item, 'direction'), get_formatted_val(item, 'lane'),
                        float(f"{item['start']:.2f}"), float(f"{item['end']:.2f}"), item['length'],
                        item.get('pavement', ''),
                        item['method'],
                        _fmt_val(item.get('score')), _fmt_val(item.get('cur_di')),
                        _fmt_val(item.get('aar_avg')), _fmt_val(item.get('iri_avg')), _fmt_val(item.get('rd_avg')),
                        item.get('reason', '')
                    ])
                wb.save(fpath)
                self._show_info("성공", "저장되었습니다.")
            except Exception as e:
                self._show_error("오류", str(e))

        self._create_button(top_frame, text="Excel 내보내기", command=on_export_priority_excel, width=120, fg_color="#2D3748").pack(side="right", padx=5)
        self._create_button(top_frame, text="알고리즘 설명", command=show_algorithm_info, width=120, fg_color="#4A5568").pack(side="right", padx=5)

        filter_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        filter_frame.pack(fill="x", padx=10, pady=(0, 5))
        var_filter = tk.StringVar()
        ent_filter = ctk.CTkEntry(filter_frame, textvariable=var_filter, width=250, placeholder_text="노선명, 방향 등 검색...")
        ent_filter.pack(side="left", padx=5)

        # 다크테마 스타일 트리뷰
        cols = ("rank", "route", "direction", "lane", "start", "end", "length", "pavement", "method", "score", "cur_di", "aar_avg", "iri_avg", "rd_avg", "reason")
        style = ttk.Style(dlg)
        style.theme_use("clam")
        style.configure("Treeview.Heading", background="#4A5568", foreground="white", relief="flat", font=(self.font_family, 10, "bold"))
        style.map("Treeview.Heading", background=[("active", "#2D3748")])
        style.configure("Treeview", background="#2D3748", fieldbackground="#2D3748", foreground="white", rowheight=25, borderwidth=0)
        style.map("Treeview", background=[("selected", "#3182CE")])

        tree = ttk.Treeview(dlg, columns=cols, show="headings", selectmode="extended")
        col_names = {"rank":"순위","route":"노선명","direction":"방향","lane":"차로",
                     "start":"시작(km)","end":"끝(km)","length":"연장(m)","pavement":"포장형식","method":"적용공법",
                     "score":"3개년DI","cur_di":"작년DI","aar_avg":"AAR","iri_avg":"IRI","rd_avg":"RD","reason":"비고"}
        for c in cols:
            tree.heading(c, text=col_names[c]); tree.column(c, anchor="center")
        tree.column("rank", width=45); tree.column("route", width=110); tree.column("direction", width=65)
        tree.column("lane", width=35); tree.column("start", width=65); tree.column("end", width=65)
        tree.column("length", width=60); tree.column("pavement", width=100); tree.column("method", width=130)
        tree.column("score", width=70); tree.column("cur_di", width=65)
        tree.column("aar_avg", width=55); tree.column("iri_avg", width=55); tree.column("rd_avg", width=55)
        tree.column("reason", width=180, anchor="w")
        tree.pack(fill="both", expand=True, padx=10, pady=10)

        last_calculated_data = []
        active_filters = {}
        tree_iid_map = {}   # tree iid(str) -> 원본 item dict

        def _fmt_dir(d):
            return d.replace("방향", "").strip()

        def _fmt_lane(l):
            return l.replace("차로", "").strip()

        def get_formatted_val(item, col):
            if col in ['score', 'cur_di', 'start', 'end']:
                v = item.get(col, 0)
                return f"{v:.2f}" if v else "-"
            if col in ['aar_avg', 'iri_avg', 'rd_avg']:
                v = item.get(col)
                return f"{v:.2f}" if v is not None else "-"
            if col == 'direction':
                return _fmt_dir(str(item.get(col, "")))
            if col == 'lane':
                return _fmt_lane(str(item.get(col, "")))
            if col == 'pavement':
                return str(item.get(col, "") or "")
            return str(item.get(col, ""))

        def build_priority_reason(is_aar, rd_severe, curr_avg, score_di, avg_aar, avg_rd, avg_iri, length_m):
            parts = [f"작년DI {curr_avg:.1f}", f"3개년DI {score_di:.1f}", f"{length_m}m"]
            if is_aar:
                parts.insert(0, f"AAR 평균 {avg_aar:.1f}로 기능성 중간층 대상")
            elif rd_severe:
                parts.insert(0, f"RD 평균 {avg_rd:.1f}로 소성변형 대응")
            else:
                parts.insert(0, "일반 절삭덧씌우기 대상")
            parts.append("최근 보수·구조물 제외 반영")
            if avg_iri is not None:
                parts.append(f"IRI {avg_iri:.1f}")
            return " | ".join(parts)

        def refresh_tree():
            for item in tree.get_children(): tree.delete(item)
            tree_iid_map.clear()
            kwd = var_filter.get().lower()
            for i, item in enumerate(last_calculated_data):
                row_vals = {c: get_formatted_val(item, c) for c in cols}
                if kwd and kwd not in " ".join(row_vals.values()).lower(): continue
                if any(row_vals[c] not in active_filters[c] for c in active_filters): continue
                tier_tag = "tier1" if item['tier'] == 1 else ("tier2" if item['tier'] == 2 else "tier3")
                iid = f"row{i}"
                tree.insert("", "end", iid=iid, values=[row_vals[c] for c in cols], tags=(tier_tag,))
                tree_iid_map[iid] = item
        
        tree.tag_configure("tier1", background="#2D3748", foreground="white")
        tree.tag_configure("tier2", background="#285E61", foreground="white")
        tree.tag_configure("tier3", background="#44337A", foreground="white")

        def calculate_priority():
            try:
                plan_year = int(var_plan_year.get())
                base_year = str(plan_year - 1)          # 전년도 (DI 기준)
                recent_years = [str(y) for y in range(plan_year - 3, plan_year)]  # 3개년
            except ValueError:
                self._show_error("오류", "연도를 확인하세요.")
                return

            candidates = []

            for route in self.routes:
                di_data    = route.get('di_data', {})
                hpci_data  = route.get('hpci_data', {})
                aar_data   = route.get('aar_data', {})
                structures = route.get('structures', [])

                # ── Step 1: 전년도 DI 맵 생성 ──────────────────────────────────
                # key: (direction, lane, km_float)  /  value: DI float
                full_di_map = {}
                dl_set = set()
                for (d, l, skm_str), val_map in di_data.items():
                    v = val_map.get(base_year)
                    if v:
                        try:
                            km_f = round(float(skm_str), 1)
                            full_di_map[(d, l, km_f)] = float(v)
                            dl_set.add((d, l))
                        except Exception:
                            pass

                for (d, l) in dl_set:

                    # ── Step 2: 방해물 수집 (구조물 + 3년 내 보수이력 모든 공법) ──
                    blockers = []

                    for s in structures:
                        try:
                            s_s, s_e = float(s['start']), float(s['end'])
                        except Exception:
                            continue
                        s_dir = s.get('direction', '양방향')
                        if s_dir != '양방향' and (s_dir not in d) and (d not in s_dir):
                            continue
                        if s_s < s_e:
                            blockers.append((s_s, s_e))

                    for entry in route.get('entries', []):
                        wd = str(entry.get("work_date", ""))
                        if len(wd) < 4:
                            continue
                        try:
                            ey = int(wd[:4])
                        except ValueError:
                            continue
                        warranty = get_method_warranty_period(entry.get("method", ""))
                        if not (plan_year - warranty <= ey <= plan_year - 1):
                            continue
                        e_dir = entry.get("direction", "")
                        if e_dir and e_dir != '양방향' and (e_dir not in d) and (d not in e_dir):
                            continue
                        if not lane_conflict(l, entry.get("lane", "전차로")):
                            continue
                        try:
                            h_s, h_e = float(entry["start"]), float(entry["end"])
                        except Exception:
                            continue
                        if h_s < h_e:
                            blockers.append((h_s, h_e))

                    # 방해물 정렬 및 겹침 병합
                    blockers.sort(key=lambda x: x[0])
                    merged = []
                    for bs, be in blockers:
                        if merged and bs < merged[-1][1]:
                            merged[-1] = (merged[-1][0], max(merged[-1][1], be))
                        else:
                            merged.append((bs, be))

                    # ── Step 3: DI 포인트 범위 파악 ──────────────────────────────
                    km_list = sorted(km for (dd, ll, km) in full_di_map if dd == d and ll == l)
                    if not km_list:
                        continue
                    route_s = km_list[0]
                    route_e = round(km_list[-1] + GRID_KM, 1)

                    # ── Step 4: 방해물로 clean 구간 분리 (연장 최대화) ───────────
                    # 방해물 사이의 공백이 clean 구간이 됨
                    clean_zones = []
                    curr = route_s
                    for bs, be in merged:
                        if bs > curr + 1e-9:
                            clean_zones.append((curr, bs))
                        curr = max(curr, be)
                    if curr < route_e - 1e-9:
                        clean_zones.append((curr, route_e))

                    # ── Step 5: 각 clean 구간 평가 ───────────────────────────────
                    for z_s, z_e in clean_zones:
                        length_m = int(round((z_e - z_s) * 1000))
                        if length_m < 300:
                            continue

                        # 구간 내 전년도 DI 포인트 수집
                        seg_pts = []  # [(km, di_val), ...]
                        k = round(math.floor(z_s / GRID_KM) * GRID_KM, 1)
                        while k < z_e - 1e-9:
                            val = full_di_map.get((d, l, round(k, 1)))
                            if val is not None:
                                seg_pts.append((round(k, 1), val))
                            k = round(k + GRID_KM, 1)

                        if not seg_pts:
                            continue

                        # 전년도 평균 DI >= 5.0 필터
                        curr_avg = sum(v for _, v in seg_pts) / len(seg_pts)
                        if curr_avg < 5.0:
                            continue

                        # ── Step 6: 3개년 DI 평균 (1차 정렬 기준) ─────────────────
                        sum_di3 = 0; cnt_di3 = 0
                        for p_km, _ in seg_pts:
                            v_map = (di_data.get((d, l, f"{p_km:.2f}")) or
                                     di_data.get((d, l, f"{p_km:.1f}")))
                            if v_map:
                                for y in recent_years:
                                    v = v_map.get(y)
                                    if v:
                                        try: sum_di3 += float(v); cnt_di3 += 1
                                        except Exception: pass
                        score_di = (sum_di3 / cnt_di3) if cnt_di3 > 0 else curr_avg

                        # ── Step 7: 3개년 HPCI 평균 (2차 정렬 기준) ───────────────
                        sum_hpci = 0; cnt_hpci = 0
                        for p_km, _ in seg_pts:
                            h_map = (hpci_data.get((d, l, f"{p_km:.2f}")) or
                                     hpci_data.get((d, l, f"{p_km:.1f}")))
                            if h_map:
                                for y in recent_years:
                                    v = h_map.get(y)
                                    if v:
                                        try: sum_hpci += float(v); cnt_hpci += 1
                                        except Exception: pass
                        score_hpci = (sum_hpci / cnt_hpci) if cnt_hpci > 0 else 0.0

                        # ── Step 8: AAR 등급 확인 (전년도 평균 >= 2.0) ────────────
                        sum_aar = 0; cnt_aar = 0
                        for p_km, _ in seg_pts:
                            a_map = (aar_data.get((d, l, f"{p_km:.2f}")) or
                                     aar_data.get((d, l, f"{p_km:.1f}")))
                            if a_map:
                                v = a_map.get(base_year)
                                if v:
                                    try: sum_aar += float(v); cnt_aar += 1
                                    except Exception: pass
                        avg_aar = (sum_aar / cnt_aar) if cnt_aar > 0 else 0.0
                        is_aar = cnt_aar > 0 and avg_aar >= 2.0

                        # ── Step 8-1: RD 등급 확인 (5등급 이하 → 12cm) ───────────
                        rd_data_r = route.get('rd_data', {})
                        sum_rd = 0; cnt_rd = 0
                        for p_km, _ in seg_pts:
                            r_map = (rd_data_r.get((d, l, f"{p_km:.2f}")) or
                                     rd_data_r.get((d, l, f"{p_km:.1f}")))
                            if r_map:
                                v = r_map.get(base_year)
                                if v:
                                    try: sum_rd += float(v); cnt_rd += 1
                                    except Exception: pass
                        avg_rd = (sum_rd / cnt_rd) if cnt_rd > 0 else 0.0
                        rd_severe = cnt_rd > 0 and avg_rd >= 5.0

                        # ── Step 8-2: IRI 평균 ────────────────────────────────────
                        iri_data_r = route.get('iri_data', {})
                        sum_iri = 0; cnt_iri = 0
                        for p_km, _ in seg_pts:
                            i_map = (iri_data_r.get((d, l, f"{p_km:.2f}")) or
                                     iri_data_r.get((d, l, f"{p_km:.1f}")))
                            if i_map:
                                v = i_map.get(base_year)
                                if v:
                                    try: sum_iri += float(v); cnt_iri += 1
                                    except Exception: pass
                        avg_iri = (sum_iri / cnt_iri) if cnt_iri > 0 else 0.0

                        # ── Step 8-3: 포장형식 (최빈값) ──────────────────────────────
                        pav_data_r = route.get('pavement_data', {})
                        pav_counter = {}
                        for p_km, _ in seg_pts:
                            pv = (pav_data_r.get((d, l, f"{p_km:.2f}")) or
                                  pav_data_r.get((d, l, f"{p_km:.1f}")))
                            if pv:
                                pav_counter[pv] = pav_counter.get(pv, 0) + 1
                        zone_pavement = max(pav_counter, key=pav_counter.get) if pav_counter else ""

                        # tier 1=일반(먼저/어두운색), tier 2=AAR(마지막/초록색)
                        tier = 2 if is_aar else 1

                        if is_aar:
                            method = "기능성 중간층"
                        elif rd_severe:
                            method = "절삭 덧씌우기(12cm)"
                        else:
                            method = "절삭 덧씌우기(10cm)"
                        reason = build_priority_reason(
                            is_aar=is_aar,
                            rd_severe=rd_severe,
                            curr_avg=curr_avg,
                            score_di=score_di,
                            avg_aar=avg_aar,
                            avg_rd=avg_rd,
                            avg_iri=avg_iri,
                            length_m=length_m,
                        )

                        candidates.append({
                            "tier":        tier,
                            "route":       route['name'],
                            "direction":   d,
                            "lane":        l,
                            "start":       z_s,
                            "end":         z_e,
                            "length":      length_m,
                            "pavement":    zone_pavement,
                            "method":      method,
                            "score":       score_di,
                            "hpci_score":  score_hpci,
                            "cur_di":      curr_avg,
                            "aar_avg":     avg_aar,
                            "iri_avg":     avg_iri,
                            "rd_avg":      avg_rd,
                            "reason":      reason,
                        })

            # ── Step 9: 우선순위 정렬 ────────────────────────────────────────────
            # 일반(tier=1) 먼저 → AAR(tier=2) 마지막
            # 동일 tier 내: ① 3년DI평균 내림차순 ② HPCI 3년평균 내림차순 ③ 연장 내림차순
            candidates.sort(key=lambda x: (
                x['tier'],
                -x['score'],
                -x['hpci_score'],
                -x['length'],
            ))

            # ── Step 10: 표면개량 구간 추가 ──────────────────────────────────────
            # 덧씌우기 구간을 방해물로 추가한 뒤, 잔여 구간 중 IRI 또는 RD 5등급 이하인 곳
            overlay_by_dl = {}  # (route_name, d, l) -> [(s, e), ...]
            for c in candidates:
                key = (c['route'], c['direction'], c['lane'])
                overlay_by_dl.setdefault(key, []).append((c['start'], c['end']))

            surface_candidates = []
            for route in self.routes:
                iri_data_r  = route.get('iri_data', {})
                rd_data_r   = route.get('rd_data', {})
                pav_data_r  = route.get('pavement_data', {})
                structures = route.get('structures', [])

                # IRI/RD 데이터가 있는 (방향, 차로) 조합 수집
                dl_iri = {(d, l) for d, l, _ in iri_data_r}
                dl_rd  = {(d, l) for d, l, _ in rd_data_r}
                all_dl = dl_iri | dl_rd

                for (d, l) in all_dl:
                    # 이 (방향,차로)의 방해물: 구조물 + 하자보수기간 내 이력 + 덧씌우기 구간
                    blockers = list(overlay_by_dl.get((route['name'], d, l), []))

                    for s in structures:
                        try:
                            s_s, s_e = float(s['start']), float(s['end'])
                        except Exception:
                            continue
                        s_dir = s.get('direction', '양방향')
                        if s_dir != '양방향' and (s_dir not in d) and (d not in s_dir):
                            continue
                        if s_s < s_e:
                            blockers.append((s_s, s_e))

                    for entry in route.get('entries', []):
                        wd = str(entry.get("work_date", ""))
                        if len(wd) < 4:
                            continue
                        try:
                            ey = int(wd[:4])
                        except ValueError:
                            continue
                        warranty = get_method_warranty_period(entry.get("method", ""))
                        if not (plan_year - warranty <= ey <= plan_year - 1):
                            continue
                        e_dir = entry.get("direction", "")
                        if e_dir and e_dir != '양방향' and (e_dir not in d) and (d not in e_dir):
                            continue
                        if not lane_conflict(l, entry.get("lane", "전차로")):
                            continue
                        try:
                            h_s, h_e = float(entry["start"]), float(entry["end"])
                        except Exception:
                            continue
                        if h_s < h_e:
                            blockers.append((h_s, h_e))

                    blockers.sort(key=lambda x: x[0])
                    merged_b = []
                    for bs, be in blockers:
                        if merged_b and bs < merged_b[-1][1]:
                            merged_b[-1] = (merged_b[-1][0], max(merged_b[-1][1], be))
                        else:
                            merged_b.append((bs, be))

                    def is_blocked(km_val):
                        for bs, be in merged_b:
                            if bs <= km_val < be:
                                return True
                        return False

                    # 전년도 IRI/RD 등급 수집 (100m 단위)
                    iri_bad = {}   # km -> True
                    for (dd, ll, skm_str), val_map in iri_data_r.items():
                        if dd != d or ll != l:
                            continue
                        v = val_map.get(base_year)
                        if v is None:
                            continue
                        try:
                            if float(v) >= 5.0:
                                iri_bad[round(float(skm_str), 1)] = True
                        except Exception:
                            pass

                    rd_bad = {}   # km -> True (아스팔트 전용 - RD 데이터 있으면 아스팔트로 간주)
                    for (dd, ll, skm_str), val_map in rd_data_r.items():
                        if dd != d or ll != l:
                            continue
                        v = val_map.get(base_year)
                        if v is None:
                            continue
                        try:
                            if float(v) >= 5.0:
                                rd_bad[round(float(skm_str), 1)] = True
                        except Exception:
                            pass

                    # 나쁜 포인트 = IRI 불량 OR RD 불량, 방해물 제외
                    bad_pts = {}  # km -> reason_set
                    for km_val in iri_bad:
                        if not is_blocked(km_val):
                            bad_pts.setdefault(km_val, set()).add("IRI")
                    for km_val in rd_bad:
                        if not is_blocked(km_val):
                            bad_pts.setdefault(km_val, set()).add("RD")

                    if not bad_pts:
                        continue

                    # 연속 구간으로 묶기 (100m 간격)
                    sorted_pts = sorted(bad_pts.keys())
                    runs = []   # [(start, end, reason_set), ...]
                    run_start = sorted_pts[0]
                    run_end   = sorted_pts[0]
                    run_reasons = set(bad_pts[sorted_pts[0]])

                    for km_val in sorted_pts[1:]:
                        if round(km_val - run_end, 1) <= GRID_KM + 1e-9:
                            run_end = km_val
                            run_reasons |= bad_pts[km_val]
                        else:
                            runs.append((run_start, round(run_end + GRID_KM, 1), set(run_reasons)))
                            run_start = km_val
                            run_end   = km_val
                            run_reasons = set(bad_pts[km_val])
                    runs.append((run_start, round(run_end + GRID_KM, 1), set(run_reasons)))

                    for r_s, r_e, reasons in runs:
                        length_m = int(round((r_e - r_s) * 1000))
                        if length_m < 100:
                            continue
                        if "IRI" in reasons and "RD" in reasons:
                            surf_reason = "표면개량 대상 | IRI·RD 동시 불량"
                        elif "IRI" in reasons:
                            surf_reason = "표면개량 대상 | IRI 불량"
                        else:
                            surf_reason = "표면개량 대상 | RD 불량"

                        # 구간 내 IRI/RD 평균 계산
                        iri_vals_zone = []
                        rd_vals_zone  = []
                        km_pt = r_s
                        while km_pt < r_e - 1e-9:
                            i_map = (iri_data_r.get((d, l, f"{km_pt:.2f}")) or
                                     iri_data_r.get((d, l, f"{km_pt:.1f}")))
                            if i_map:
                                iri_v = i_map.get(base_year)
                                if iri_v is not None:
                                    try: iri_vals_zone.append(float(iri_v))
                                    except Exception: pass
                            r_map = (rd_data_r.get((d, l, f"{km_pt:.2f}")) or
                                     rd_data_r.get((d, l, f"{km_pt:.1f}")))
                            if r_map:
                                rd_v = r_map.get(base_year)
                                if rd_v is not None:
                                    try: rd_vals_zone.append(float(rd_v))
                                    except Exception: pass
                            km_pt = round(km_pt + GRID_KM, 1)
                        avg_iri_s = (sum(iri_vals_zone) / len(iri_vals_zone)) if iri_vals_zone else None
                        avg_rd_s  = (sum(rd_vals_zone)  / len(rd_vals_zone))  if rd_vals_zone  else None
                        iri_disp = f"{avg_iri_s:.1f}" if avg_iri_s is not None else "-"
                        rd_disp = f"{avg_rd_s:.1f}" if avg_rd_s is not None else "-"

                        # 포장형식 (최빈값)
                        pav_counter_s = {}
                        km_pt2 = r_s
                        while km_pt2 < r_e - 1e-9:
                            pv = (pav_data_r.get((d, l, f"{km_pt2:.2f}")) or
                                  pav_data_r.get((d, l, f"{km_pt2:.1f}")))
                            if pv:
                                pav_counter_s[pv] = pav_counter_s.get(pv, 0) + 1
                            km_pt2 = round(km_pt2 + GRID_KM, 1)
                        zone_pav_s = max(pav_counter_s, key=pav_counter_s.get) if pav_counter_s else ""

                        surface_candidates.append({
                            "tier":        3,
                            "route":       route['name'],
                            "direction":   d,
                            "lane":        l,
                            "start":       r_s,
                            "end":         r_e,
                            "length":      length_m,
                            "pavement":    zone_pav_s,
                            "method":      "표면개량",
                            "score":       0.0,
                            "hpci_score":  0.0,
                            "cur_di":      None,
                            "aar_avg":     None,
                            "iri_avg":     avg_iri_s,
                            "rd_avg":      avg_rd_s,
                            "reason":      f"{surf_reason} | 연장 {length_m}m | IRI {iri_disp} | RD {rd_disp}",
                        })

            surface_candidates.sort(key=lambda x: -x['length'])
            all_results = candidates + surface_candidates
            for i, c in enumerate(all_results):
                c['rank'] = i + 1

            # 운영계획변경 등에서 특정 구간(사업계획 지정 구간)을 후보에서 제외
            if exclude_sections:
                def _is_excluded(r):
                    for ex in exclude_sections:
                        if (str(r.get('route')) == str(ex.get('route'))
                                and str(r.get('direction')) == str(ex.get('direction'))
                                and str(r.get('lane')) == str(ex.get('lane'))
                                and intervals_overlap(float(r.get('start', 0)), float(r.get('end', 0)),
                                                      float(ex.get('start', 0)), float(ex.get('end', 0)))):
                            return True
                    return False
                all_results = [r for r in all_results if not _is_excluded(r)]
                for i, c in enumerate(all_results):
                    c['rank'] = i + 1

            nonlocal last_calculated_data
            last_calculated_data = all_results
            active_filters.clear()
            refresh_tree()
            self._show_priority_complete_popup("완료",
                f"총 {len(all_results)}개 구간 선정됨.",
                parent=dlg
            )

        btn_calc = self._create_button(top_frame, text="우선순위 산정", command=calculate_priority, width=120, font=(self.font_family, 12, "bold"), fg_color="#3182CE")
        btn_calc.pack(side="left", padx=10)

        # 선택 적용 버튼 (사업계획/운영계획변경에서 호출 시)
        if apply_label and callable(apply_callback):
            def _on_apply_selection():
                sel = tree.selection()
                items = [tree_iid_map[i] for i in sel if i in tree_iid_map]
                if not items:
                    self._show_info("알림", "적용할 구간을 선택해 주세요.")
                    return
                # 닫기 전에 선택 항목을 깊은 복사하여 전달 (원본 변형 방지)
                payload = [dict(it) for it in items]
                plan_year = var_plan_year.get()
                try:
                    dlg.destroy()
                except Exception:
                    pass
                self.priority_window = None
                apply_callback(payload, plan_year)
            self._create_button(top_frame, text=apply_label, command=_on_apply_selection,
                                width=140, font=(self.font_family, 12, "bold"),
                                fg_color=PRIMARY_BLUE, hover_color=PRIMARY_BLUE_HOVER,
                                text_color="#FFFFFF").pack(side="left", padx=6)
        btn_search = self._create_button(filter_frame, text="검색", command=refresh_tree, width=80)
        btn_search.pack(side="left", padx=5)
        ent_filter.bind("<Return>", lambda e: refresh_tree())

        def sort_data(col, reverse):
            try:
                if col in ['rank', 'start', 'end', 'length', 'score', 'current_avg']:
                    last_calculated_data.sort(key=lambda x: float(x.get(col, 0)), reverse=reverse)
                else: last_calculated_data.sort(key=lambda x: str(x.get(col, "")), reverse=reverse)
            except Exception: last_calculated_data.sort(key=lambda x: str(x.get(col, "")), reverse=reverse)
            refresh_tree()

        def open_header_menu(col_name, event):
            unique_vals = sorted(list(set(get_formatted_val(i, col_name) for i in last_calculated_data)))
            if not unique_vals: return
            
            fdlg = self._create_popup_window(dlg)
            fdlg.wm_overrideredirect(True)
            fdlg.attributes('-topmost', True)
            
            w, h = 260, 420
            x, y = event.x_root, event.y_root
            try:
                if x + w > fdlg.winfo_screenwidth(): x = fdlg.winfo_screenwidth() - w
                if y + h > fdlg.winfo_screenheight(): y = fdlg.winfo_screenheight() - h
            except Exception: pass
            fdlg.geometry(f"{w}x{h}+{x}+{y}")
            
            fdlg.lift()
            fdlg.focus_force()

            closing = {"done": False}
            outside_click_bind_id = None

            def close_menu():
                if closing["done"]:
                    return
                closing["done"] = True
                try:
                    if outside_click_bind_id:
                        dlg.unbind("<Button-1>", outside_click_bind_id)
                except Exception:
                    pass
                try:
                    fdlg.destroy()
                except Exception:
                    pass

            def check_outside_click(e):
                try:
                    widget = e.widget
                    if widget and widget.winfo_toplevel() == fdlg:
                        return
                    cx, cy = e.x_root, e.y_root
                    rx, ry = fdlg.winfo_rootx(), fdlg.winfo_rooty()
                    rw, rh = fdlg.winfo_width(), fdlg.winfo_height()
                    if not (rx <= cx <= rx + rw and ry <= cy <= ry + rh):
                        close_menu()
                except Exception:
                    close_menu()

            outside_click_bind_id = dlg.bind("<Button-1>", check_outside_click, add="+")
            fdlg.bind("<FocusOut>", lambda _e: close_menu())
            fdlg.bind("<Escape>", lambda _e: close_menu())
            fdlg.protocol("WM_DELETE_WINDOW", close_menu)
            
            mf = ctk.CTkFrame(fdlg, fg_color="#FFFFFF", border_width=1, border_color="#C8D7EA", corner_radius=12)
            mf.pack(fill="both", expand=True)
            
            def srt(rev): 
                sort_data(col_name, rev)
                close_menu()
                
            self._create_button(
                mf, text="오름차순", command=lambda:srt(False),
                fg_color="transparent", anchor="w", text_color=TITLE_TEXT,
                hover_color="#E8F1FB", corner_radius=8, height=30
            ).pack(fill="x", padx=6, pady=(6, 2))
            self._create_button(
                mf, text="내림차순", command=lambda:srt(True),
                fg_color="transparent", anchor="w", text_color=TITLE_TEXT,
                hover_color="#E8F1FB", corner_radius=8, height=30
            ).pack(fill="x", padx=6)
            ctk.CTkFrame(mf, height=1, fg_color="#E2EAF4").pack(fill="x", padx=6, pady=6)
            
            sf = ctk.CTkScrollableFrame(mf, fg_color="#F8FBFF", corner_radius=10)
            sf.pack(fill="both", expand=True, padx=6, pady=(0, 6))
            
            cur_set = active_filters.get(col_name, set(unique_vals))
            v_map = {}
            
            def apply():
                active_filters[col_name] = {v for v,var in v_map.items() if var.get()}
                if len(active_filters[col_name]) == len(unique_vals): del active_filters[col_name]
                refresh_tree()
            
            all_v = tk.BooleanVar(value=(len(cur_set)==len(unique_vals)))
            def toggle_all():
                val = all_v.get()
                for v in v_map.values(): v.set(val)
                apply()

            ctk.CTkCheckBox(
                sf, text="(전체)", variable=all_v, command=toggle_all,
                text_color=TITLE_TEXT, checkbox_width=18, checkbox_height=18,
                border_width=2, fg_color=PRIMARY_BLUE, hover_color=PRIMARY_BLUE_HOVER,
                border_color="#AFC4DE"
            ).pack(anchor="w", pady=2)
            
            for v in unique_vals:
                var = tk.BooleanVar(value=(v in cur_set))
                v_map[v] = var
                ctk.CTkCheckBox(
                    sf, text=str(v), variable=var, command=apply,
                    text_color=TITLE_TEXT, checkbox_width=18, checkbox_height=18,
                    border_width=2, fg_color=PRIMARY_BLUE, hover_color=PRIMARY_BLUE_HOVER,
                    border_color="#AFC4DE"
                ).pack(anchor="w", pady=1)
                
            self._create_button(
                mf, text="닫기", command=close_menu, height=28,
                fg_color="#EEF4FB", hover_color="#E0ECF8", text_color=TITLE_TEXT,
                border_width=1, border_color="#C8D7EA"
            ).pack(fill="x", padx=6, pady=(0, 6))

        def on_head_click(e):
            r = tree.identify_region(e.x, e.y)
            if r == "heading":
                cid = tree.identify_column(e.x)
                if cid: open_header_menu(cols[int(cid[1:])-1], e)
        tree.bind("<Button-1>", on_head_click)

        def on_priority_row_dblclick(e):
            sel = tree.selection()
            if not sel: return
            vals = tree.item(sel[0], "values")
            try:
                rname = str(vals[1])
                direction = str(vals[2])
                lane_raw  = str(vals[3]).strip()
                lane = lane_raw + "차로" if lane_raw and not lane_raw.endswith("차로") else lane_raw
                s = float(str(vals[4]).replace(",", ""))
                en = float(str(vals[5]).replace(",", ""))
                self.navigate_to_section(rname, s, en, direction=direction, lane=lane)
            except Exception:
                pass
        tree.bind("<Double-1>", on_priority_row_dblclick)

    def on_defect_risk(self):
        """하자발생 우려구간 분석 창을 엽니다."""
        if hasattr(self, "_defect_window") and self._defect_window is not None and self._defect_window.winfo_exists():
            self._defect_window.lift()
            self._defect_window.focus_force()
            return

        dlg = self._create_popup_window(self)
        self._defect_window = dlg
        dlg.title("")
        dlg.geometry("1200x600")
        dlg.transient(self)
        dlg.lift()
        dlg.focus_force()

        # 상단 버튼 행
        top = ctk.CTkFrame(dlg, fg_color="transparent")
        top.pack(fill="x", padx=10, pady=10)

        defect_data = []

        lbl_count = ctk.CTkLabel(top, text="", font=(self.font_family, 13, "bold"), text_color="#EF4444")
        lbl_count.pack(side="right", padx=(0, 10))

        def show_algorithm_info():
            info = self._create_popup_window(dlg)
            info.title("")
            info.geometry("720x480")
            info.transient(dlg)
            info.grab_set()
            info.lift()
            info.focus_force()
            txt = ctk.CTkTextbox(info, font=(self.font_family, 12), wrap="word")
            txt.pack(fill="both", expand=True, padx=16, pady=(14, 8))
            txt.insert("end", (
                "■ 하자발생 우려구간 선정 알고리즘\n\n"
                "1. [검사 대상 이력 필터]\n"
                "   - 보수 시공일로부터 현재까지 최소 2년 이상 경과한 이력만 대상\n"
                "   - 구간 연장이 200m(0.2km) 미만이면 제외\n\n"
                "2. [유효 검사 구간 설정]\n"
                "   - 보수 구간 앞뒤 각 100m를 제외한 내부 구간만 검사\n"
                "   - 예) 3.5~3.9km → 3.6~3.8km 구간 내 셀만 검사\n\n"
                "3. [포장상태 판단 지표 선택]\n"
                "   ─ 일반 공법 (덧씌우기 등)\n"
                "     · DI 지수 5.0 이상인 100m 셀이 하나라도 있으면 하자 우려\n"
                "   ─ 표면개량 카테고리 공법\n"
                "     · RD(소성변형) ≥ 5.0 또는 IRI(평탄성) ≥ 3.5 인 셀 존재 시 하자 우려\n\n"
                "4. [검사 연도 범위]\n"
                "   - 보수연도 기준 +2년 ~ +4년의 포장상태 데이터를 확인\n"
                "   - 예) 2022년 보수 → 2024, 2025, 2026년 데이터 검토\n\n"
                "5. [하자발생우려근거 표기]\n"
                "   - 불량 셀 중 상위 2개 위치·수치·연도를 근거에 표시\n"
                "   - 표면개량 공법은 RD/IRI 각각 구분하여 병기\n"
            ))
            txt.configure(state="disabled")
            self._create_close_button(info, info.destroy, width=100, height=32).pack(pady=(0, 12))

        ctk.CTkLabel(top, text="하자발생 우려구간 분석",
                     font=(self.font_family, 14, "bold")).pack(side="left")

        # 트리뷰
        style = ttk.Style(dlg)
        style.theme_use("clam")
        style.configure("Defect.Treeview.Heading", background="#4A5568", foreground="white",
                        relief="flat", font=(self.font_family, 10, "bold"))
        style.map("Defect.Treeview.Heading", background=[("active", "#2D3748")])
        style.configure("Defect.Treeview", background="#2D3748", fieldbackground="#2D3748",
                        foreground="white", rowheight=26, borderwidth=0)
        style.map("Defect.Treeview", background=[("selected", "#7B2222")])

        d_cols = ("route", "section", "direction", "method", "work_year", "reason")
        d_names = {
            "route":     "노선",
            "section":   "이정(km)",
            "direction": "방향",
            "method":    "적용공법",
            "work_year": "보수연도",
            "reason":    "하자발생우려근거",
        }

        btn_row = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_row.pack(fill="x", padx=10, pady=(0, 4))

        tree_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        tree_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        defect_tree = ttk.Treeview(tree_frame, columns=d_cols, show="headings",
                                   selectmode="browse", style="Defect.Treeview")
        for c in d_cols:
            defect_tree.heading(c, text=d_names[c])
            defect_tree.column(c, anchor="center")
        defect_tree.column("route",     width=110)
        defect_tree.column("section",   width=155)
        defect_tree.column("direction", width=80)
        defect_tree.column("method",    width=135)
        defect_tree.column("work_year", width=70)
        defect_tree.column("reason",    width=560, anchor="w")
        defect_tree.tag_configure("risk")  # 별도 배경색 없음

        sb = ctk.CTkScrollbar(tree_frame, orientation="vertical", command=defect_tree.yview)
        defect_tree.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        defect_tree.pack(side="left", fill="both", expand=True)

        def on_defect_row_dblclick(e):
            sel = defect_tree.selection()
            if not sel: return
            vals = defect_tree.item(sel[0], "values")
            try:
                rname = str(vals[0])
                parts = str(vals[1]).split("~")
                s = float(parts[0])
                en = float(parts[1])
                direction = str(vals[2]) if len(vals) > 2 else None
                # defect_data에서 해당 행의 lane 조회
                lane = None
                row_idx = defect_tree.index(sel[0])
                if 0 <= row_idx < len(defect_data):
                    lane = defect_data[row_idx].get("lane")
                self.navigate_to_section(rname, s, en, direction=direction, lane=lane)
            except Exception:
                pass
        defect_tree.bind("<Double-1>", on_defect_row_dblclick)

        def calculate_defects():
            nonlocal defect_data
            defect_data = []
            cur_year = datetime.now().year

            for route in self.routes:
                rname     = route.get("name", "")
                di_data   = route.get("di_data",   {})
                rd_data   = route.get("rd_data",   {})
                iri_data  = route.get("iri_data",  {})
                entries   = route.get("entries", [])

                for entry in entries:
                    wd = str(entry.get("work_date", ""))
                    if len(wd) < 4:
                        continue
                    try:
                        work_year = int(wd[:4])
                    except ValueError:
                        continue

                    if cur_year <= work_year:
                        continue  # 시공 후 1년도 경과하지 않음

                    e_start   = entry.get("start", 0.0)
                    e_end     = entry.get("end",   0.0)
                    direction = entry.get("direction", "")
                    method    = entry.get("method",    "")

                    # 공법 카테고리 및 하자보증기간 결정
                    method_cat = METHOD_CATEGORY_MAP.get(method, "")
                    warranty_period = get_method_warranty_period(method)

                    # 하자보증기간이 이미 만료된 경우 제외
                    if cur_year > work_year + warranty_period:
                        continue

                    # 앞뒤 100m 제외한 유효 검사 구간
                    eff_start = e_start + GRID_KM
                    eff_end   = e_end   - GRID_KM
                    if eff_end <= eff_start + 1e-9:
                        continue  # 구간이 200m 미만이면 검사 대상 없음

                    is_surface = (method_cat == "표면개량")
                    if is_surface:
                        # 표면개량 계열: RD(≥5) 또는 IRI(≥3.5) 기준
                        check_sets = [
                            (rd_data,  "RD",  5.0),
                            (iri_data, "IRI", 3.5),
                        ]
                    else:
                        # 그 외: DI 지수 5.0 이상
                        check_sets = [(di_data, "DI", 5.0)]

                    bad_cells = []
                    for data_dict, indicator, threshold in check_sets:
                        for (d, l, skm_str), year_map in data_dict.items():
                            if direction and d != direction:
                                continue
                            try:
                                cell_s = float(skm_str)
                            except (ValueError, TypeError):
                                continue
                            cell_e = cell_s + GRID_KM
                            # 앞뒤 100m 제외한 구간 안에 셀이 완전히 포함될 때만 검사
                            if cell_s < eff_start - 1e-9 or cell_e > eff_end + 1e-9:
                                continue
                            chk_end = min(work_year + warranty_period, cur_year)
                            for chk_yr in range(work_year + 2, chk_end + 1):
                                val = year_map.get(str(chk_yr))
                                if val is None:
                                    continue
                                try:
                                    fval = float(val)
                                    if fval >= threshold:
                                        bad_cells.append((cell_s, cell_e, fval, chk_yr, indicator))
                                        break
                                except (ValueError, TypeError):
                                    pass

                    if not bad_cells:
                        continue

                    worst = sorted(bad_cells, key=lambda c: c[2], reverse=True)[:3]
                    # 지표별로 그룹화해서 근거 문자열 생성
                    by_indicator: dict = {}
                    for c in bad_cells:
                        by_indicator.setdefault(c[4], []).append(c)
                    reason_parts = []
                    for ind, cells in by_indicator.items():
                        best = sorted(cells, key=lambda c: c[2], reverse=True)[:2]
                        locs = [f"{c[0]:.1f}~{c[1]:.1f}km({c[2]:.1f}/{c[3]}년)" for c in best]
                        suffix = "..." if len(cells) > 2 else ""
                        reason_parts.append(f"{ind} 불량 ({', '.join(locs)}{suffix})")
                    reason = "포장상태불량률 악화 — " + " | ".join(reason_parts)

                    defect_data.append({
                        "route":     rname,
                        "section":   f"{e_start:.3f}~{e_end:.3f}",
                        "direction": direction,
                        "lane":      entry.get("lane", "전차로"),
                        "method":    method,
                        "work_year": str(work_year),
                        "reason":    reason,
                    })

            for row in defect_tree.get_children():
                defect_tree.delete(row)
            for item in defect_data:
                defect_tree.insert("", "end", tags=("risk",), values=(
                    item["route"], item["section"], item["direction"],
                    item["method"], item["work_year"], item["reason"]
                ))
            cnt = len(defect_data)
            lbl_count.configure(text=f"총 {cnt}건 검출" if cnt else "")
            if cnt == 0:
                self._show_info("분석 완료", "하자발생 우려구간이 없습니다.")

        def export_excel():
            if not _EXCEL_LIBS_AVAILABLE:
                self._show_error("라이브러리 필요", "openpyxl 필요")
                return
            if not defect_data:
                self._show_info("알림", "먼저 분석을 실행해 주세요.")
                return
            fpath = filedialog.asksaveasfilename(
                title="하자 우려구간 저장", defaultextension=".xlsx",
                filetypes=[("Excel 파일", "*.xlsx")],
                initialfile="하자발생우려구간.xlsx", parent=dlg
            )
            if not fpath:
                return
            try:
                from openpyxl.styles import PatternFill
                wb = Workbook(); ws = wb.active; ws.title = "하자우려구간"
                hdrs = ["노선", "이정(km)", "방향", "적용공법", "보수연도", "하자발생우려근거"]
                ws.append(hdrs)
                hfont = Font(bold=True, color="FFFFFF")
                hfill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
                for cell in ws[1]:
                    cell.font = hfont; cell.fill = hfill; cell.alignment = Alignment(horizontal="center")
                for d in defect_data:
                    ws.append([d["route"], d["section"], d["direction"],
                               d["method"], d["work_year"], d["reason"]])
                for col in ws.columns:
                    mx = max((len(str(c.value or "")) for c in col), default=8)
                    ws.column_dimensions[col[0].column_letter].width = min(mx + 4, 60)
                wb.save(fpath)
                self._show_info("저장 완료", f"저장되었습니다:\n{fpath}")
            except Exception as ex:
                self._show_error("오류", str(ex))

        self._create_button(btn_row, text="알고리즘 설명", command=show_algorithm_info,
                      fg_color="#4A5568", width=120).pack(side="left", padx=(0, 8))
        self._create_button(btn_row, text="Excel 내보내기", command=export_excel,
                      fg_color="#2D3748", width=120).pack(side="left")

        # 창 열릴 때 자동 분석
        dlg.after(100, calculate_defects)

    def on_defect_inspection(self):
        """하자점검 창을 엽니다.

        현재 연도를 기준으로, 각 보수이력의 공법에 설정된 하자기간이 아직
        만료되지 않은(=하자담보책임기간 내) 구간을 표출합니다.
        금년에 하자기간이 만료되는 구간(잔여 0년)을 최상단에 두어,
        만료가 임박한 구간부터 보이도록 정렬합니다.
        노선/이정/방향/공법 등 컬럼별 필터를 제공합니다.
        """
        if hasattr(self, "_defect_inspect_window") and self._defect_inspect_window is not None \
                and self._defect_inspect_window.winfo_exists():
            self._defect_inspect_window.lift()
            self._defect_inspect_window.focus_force()
            return

        dlg = self._create_popup_window(self)
        self._defect_inspect_window = dlg
        dlg.title("")
        dlg.geometry("1280x640")
        dlg.transient(self)
        dlg.lift()
        dlg.focus_force()

        cur_year = datetime.now().year

        # 상단 제목/건수 행
        top = ctk.CTkFrame(dlg, fg_color="transparent")
        top.pack(fill="x", padx=10, pady=10)
        ctk.CTkLabel(top, text=f"하자점검 (기준연도 {cur_year}년)",
                     font=(self.font_family, 14, "bold")).pack(side="left")
        lbl_count = ctk.CTkLabel(top, text="", font=(self.font_family, 13, "bold"),
                                 text_color="#E33E4B")
        lbl_count.pack(side="right", padx=(0, 10))

        inspect_data = []
        active_filters = {}

        def show_algorithm_info():
            info = self._create_popup_window(dlg)
            info.title("")
            info.geometry("720x420")
            info.transient(dlg)
            info.grab_set()
            info.lift()
            info.focus_force()
            txt = ctk.CTkTextbox(info, font=(self.font_family, 12), wrap="word")
            txt.pack(fill="both", expand=True, padx=16, pady=(14, 8))
            txt.insert("end", (
                "■ 하자점검 대상 선정 기준\n\n"
                f"1. [기준 연도]\n"
                f"   - 현재 연도({cur_year}년)를 기준으로 판단합니다.\n\n"
                "2. [하자기간 내 구간 표출]\n"
                "   - 각 보수이력의 공법에 설정된 하자담보책임기간을 확인\n"
                "   - 보수연도 + 하자기간(만료연도)이 기준연도 이상인 구간만 표출\n"
                "   - 즉, 현재 하자담보책임기간 내에 있는 구간만 대상\n\n"
                "3. [정렬]\n"
                "   - 만료연도가 빠른(임박한) 구간을 최상단에 배치\n"
                "   - 금년에 만료되는 구간(잔여 0년)이 가장 위에 표시\n\n"
                "4. [필터]\n"
                "   - 컬럼 머리글 클릭 시 노선·이정·방향·공법 등으로 필터 가능\n"
                "   - 상단 검색창으로 키워드 검색 가능\n"
            ))
            txt.configure(state="disabled")
            self._create_close_button(info, info.destroy, width=100, height=32).pack(pady=(0, 12))

        # 필터(검색) 행
        filter_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        filter_frame.pack(fill="x", padx=10, pady=(0, 5))
        var_filter = tk.StringVar()
        ent_filter = ctk.CTkEntry(filter_frame, textvariable=var_filter, width=250,
                                  placeholder_text="노선, 이정, 공법 등 검색...")
        ent_filter.pack(side="left", padx=5)

        # 트리뷰 스타일
        style = ttk.Style(dlg)
        style.theme_use("clam")
        style.configure("Inspect.Treeview.Heading", background="#4A5568", foreground="white",
                        relief="flat", font=(self.font_family, 10, "bold"))
        style.map("Inspect.Treeview.Heading", background=[("active", "#2D3748")])
        style.configure("Inspect.Treeview", background="#2D3748", fieldbackground="#2D3748",
                        foreground="white", rowheight=26, borderwidth=0)
        style.map("Inspect.Treeview", background=[("selected", "#7B2222")])

        i_cols = ("route", "section", "direction", "lane", "method",
                  "category", "warranty", "work_year", "expire_year", "remaining")
        i_names = {
            "route":       "노선",
            "section":     "이정(km)",
            "direction":   "방향",
            "lane":        "차로",
            "method":      "적용공법",
            "category":    "하자카테고리",
            "warranty":    "하자기간(년)",
            "work_year":   "보수연도",
            "expire_year": "만료연도",
            "remaining":   "잔여(년)",
        }

        tree_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        tree_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        tree = ttk.Treeview(tree_frame, columns=i_cols, show="headings",
                            selectmode="browse", style="Inspect.Treeview")
        for c in i_cols:
            tree.heading(c, text=i_names[c])
            tree.column(c, anchor="center")
        tree.column("route",       width=120)
        tree.column("section",     width=150)
        tree.column("direction",   width=80)
        tree.column("lane",        width=60)
        tree.column("method",      width=140)
        tree.column("category",    width=110)
        tree.column("warranty",    width=85)
        tree.column("work_year",   width=75)
        tree.column("expire_year", width=75)
        tree.column("remaining",   width=70)
        # 금년 만료(잔여 0년) 강조
        tree.tag_configure("expire_now", background="#7B2222", foreground="#FFFFFF")
        tree.tag_configure("normal")

        sb = ctk.CTkScrollbar(tree_frame, orientation="vertical", command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        tree.pack(side="left", fill="both", expand=True)

        def get_val(item, col):
            return str(item.get(col, ""))

        def refresh_tree():
            for row in tree.get_children():
                tree.delete(row)
            kwd = var_filter.get().lower()
            shown = 0
            for item in inspect_data:
                row_vals = {c: get_val(item, c) for c in i_cols}
                if kwd and kwd not in " ".join(row_vals.values()).lower():
                    continue
                if any(row_vals[c] not in active_filters[c] for c in active_filters):
                    continue
                tag = "expire_now" if item.get("remaining") == 0 else "normal"
                tree.insert("", "end", values=[row_vals[c] for c in i_cols], tags=(tag,))
                shown += 1
            lbl_count.configure(text=f"총 {shown}건 (하자기간 내)" if shown else "")

        def calculate_inspection():
            nonlocal inspect_data
            inspect_data = []
            for route in self.routes:
                rname   = route.get("name", "")
                entries = route.get("entries", [])
                for entry in entries:
                    wd = str(entry.get("work_date", ""))
                    if len(wd) < 4:
                        continue
                    try:
                        work_year = int(wd[:4])
                    except ValueError:
                        continue
                    if work_year > cur_year:
                        continue  # 미래 시공(계획 등) 제외

                    method   = entry.get("method", "")
                    warranty = get_method_warranty_period(method)
                    expire_year = work_year + warranty
                    # 하자담보책임기간이 이미 만료된 구간은 제외
                    if expire_year < cur_year:
                        continue
                    remaining = expire_year - cur_year

                    try:
                        e_start = float(entry.get("start", 0.0))
                        e_end   = float(entry.get("end", 0.0))
                    except (ValueError, TypeError):
                        e_start, e_end = 0.0, 0.0

                    inspect_data.append({
                        "route":       rname,
                        "section":     f"{e_start:.3f}~{e_end:.3f}",
                        "direction":   entry.get("direction", ""),
                        "lane":        entry.get("lane", "전차로"),
                        "method":      method,
                        "category":    METHOD_CATEGORY_MAP.get(method, "") or "-",
                        "warranty":    warranty,
                        "work_year":   work_year,
                        "expire_year": expire_year,
                        "remaining":   remaining,
                        "_start":      e_start,
                        "_end":        e_end,
                    })

            # 만료가 임박한 순(만료연도 오름차순 → 잔여 오름차순), 동률은 노선/이정 순
            inspect_data.sort(key=lambda x: (x["expire_year"], x["remaining"],
                                             x["route"], x["_start"]))
            active_filters.clear()
            refresh_tree()
            if not inspect_data:
                self._show_info("하자점검", "하자기간 내 구간이 없습니다.")

        def open_header_menu(col_name, event):
            unique_vals = sorted(set(get_val(i, col_name) for i in inspect_data))
            if not unique_vals:
                return
            fdlg = self._create_popup_window(dlg)
            fdlg.wm_overrideredirect(True)
            fdlg.attributes('-topmost', True)
            w, h = 260, 420
            x, y = event.x_root, event.y_root
            try:
                if x + w > fdlg.winfo_screenwidth(): x = fdlg.winfo_screenwidth() - w
                if y + h > fdlg.winfo_screenheight(): y = fdlg.winfo_screenheight() - h
            except Exception:
                pass
            fdlg.geometry(f"{w}x{h}+{x}+{y}")
            fdlg.lift()
            fdlg.focus_force()

            closing = {"done": False}
            outside_click_bind_id = None

            def close_menu():
                if closing["done"]:
                    return
                closing["done"] = True
                try:
                    if outside_click_bind_id:
                        dlg.unbind("<Button-1>", outside_click_bind_id)
                except Exception:
                    pass
                try:
                    fdlg.destroy()
                except Exception:
                    pass

            def check_outside_click(e):
                try:
                    widget = e.widget
                    if widget and widget.winfo_toplevel() == fdlg:
                        return
                    cx, cy = e.x_root, e.y_root
                    rx, ry = fdlg.winfo_rootx(), fdlg.winfo_rooty()
                    rw, rh = fdlg.winfo_width(), fdlg.winfo_height()
                    if not (rx <= cx <= rx + rw and ry <= cy <= ry + rh):
                        close_menu()
                except Exception:
                    close_menu()

            outside_click_bind_id = dlg.bind("<Button-1>", check_outside_click, add="+")
            fdlg.bind("<FocusOut>", lambda _e: close_menu())
            fdlg.bind("<Escape>", lambda _e: close_menu())
            fdlg.protocol("WM_DELETE_WINDOW", close_menu)

            mf = ctk.CTkFrame(fdlg, fg_color="#FFFFFF", border_width=1,
                              border_color="#C8D7EA", corner_radius=12)
            mf.pack(fill="both", expand=True)

            sf = ctk.CTkScrollableFrame(mf, fg_color="#F8FBFF", corner_radius=10)
            sf.pack(fill="both", expand=True, padx=6, pady=6)

            cur_set = active_filters.get(col_name, set(unique_vals))
            v_map = {}

            def apply():
                active_filters[col_name] = {v for v, var in v_map.items() if var.get()}
                if len(active_filters[col_name]) == len(unique_vals):
                    del active_filters[col_name]
                refresh_tree()

            all_v = tk.BooleanVar(value=(len(cur_set) == len(unique_vals)))

            def toggle_all():
                val = all_v.get()
                for v in v_map.values():
                    v.set(val)
                apply()

            ctk.CTkCheckBox(
                sf, text="(전체)", variable=all_v, command=toggle_all,
                text_color=TITLE_TEXT, checkbox_width=18, checkbox_height=18,
                border_width=2, fg_color=PRIMARY_BLUE, hover_color=PRIMARY_BLUE_HOVER,
                border_color="#AFC4DE"
            ).pack(anchor="w", pady=2)

            for v in unique_vals:
                var = tk.BooleanVar(value=(v in cur_set))
                v_map[v] = var
                ctk.CTkCheckBox(
                    sf, text=str(v), variable=var, command=apply,
                    text_color=TITLE_TEXT, checkbox_width=18, checkbox_height=18,
                    border_width=2, fg_color=PRIMARY_BLUE, hover_color=PRIMARY_BLUE_HOVER,
                    border_color="#AFC4DE"
                ).pack(anchor="w", pady=1)

            self._create_button(
                mf, text="닫기", command=close_menu, height=28,
                fg_color="#EEF4FB", hover_color="#E0ECF8", text_color=TITLE_TEXT,
                border_width=1, border_color="#C8D7EA"
            ).pack(fill="x", padx=6, pady=(0, 6))

        def on_head_click(e):
            r = tree.identify_region(e.x, e.y)
            if r == "heading":
                cid = tree.identify_column(e.x)
                if cid:
                    open_header_menu(i_cols[int(cid[1:]) - 1], e)
        tree.bind("<Button-1>", on_head_click)

        def on_row_dblclick(e):
            sel = tree.selection()
            if not sel:
                return
            vals = tree.item(sel[0], "values")
            try:
                rname = str(vals[0])
                parts = str(vals[1]).split("~")
                s = float(parts[0])
                en = float(parts[1])
                direction = str(vals[2]) if len(vals) > 2 else None
                lane = str(vals[3]) if len(vals) > 3 else None
                self.navigate_to_section(rname, s, en, direction=direction, lane=lane)
            except Exception:
                pass
        tree.bind("<Double-1>", on_row_dblclick)

        def export_excel():
            if not _EXCEL_LIBS_AVAILABLE:
                self._show_error("라이브러리 필요", "openpyxl 필요")
                return
            if not inspect_data:
                self._show_info("알림", "먼저 분석을 실행해 주세요.")
                return
            fpath = filedialog.asksaveasfilename(
                title="하자점검 저장", defaultextension=".xlsx",
                filetypes=[("Excel 파일", "*.xlsx")],
                initialfile=f"{cur_year}년_하자점검.xlsx", parent=dlg
            )
            if not fpath:
                return
            try:
                from openpyxl.styles import PatternFill
                wb = Workbook(); ws = wb.active; ws.title = "하자점검"
                hdrs = [i_names[c] for c in i_cols]
                ws.append(hdrs)
                hfont = Font(bold=True, color="FFFFFF")
                hfill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
                for cell in ws[1]:
                    cell.font = hfont; cell.fill = hfill
                    cell.alignment = Alignment(horizontal="center")
                kwd = var_filter.get().lower()
                for item in inspect_data:
                    row_vals = {c: get_val(item, c) for c in i_cols}
                    if kwd and kwd not in " ".join(row_vals.values()).lower():
                        continue
                    if any(row_vals[c] not in active_filters[c] for c in active_filters):
                        continue
                    ws.append([row_vals[c] for c in i_cols])
                for col in ws.columns:
                    mx = max((len(str(c.value or "")) for c in col), default=8)
                    ws.column_dimensions[col[0].column_letter].width = min(mx + 4, 40)
                wb.save(fpath)
                self._show_info("저장 완료", f"저장되었습니다:\n{fpath}")
            except Exception as ex:
                self._show_error("오류", str(ex))

        btn_row = ctk.CTkFrame(filter_frame, fg_color="transparent")
        btn_row.pack(side="right")
        self._create_button(btn_row, text="알고리즘 설명", command=show_algorithm_info,
                            fg_color="#4A5568", width=120).pack(side="left", padx=(0, 8))
        self._create_button(btn_row, text="Excel 내보내기", command=export_excel,
                            fg_color="#2D3748", width=120).pack(side="left", padx=(0, 8))
        self._create_button(filter_frame, text="검색", command=refresh_tree, width=80).pack(side="left", padx=5)
        ent_filter.bind("<Return>", lambda e: refresh_tree())

        # 창 열릴 때 자동 분석
        dlg.after(100, calculate_inspection)

    def navigate_to_section(self, route_name: str, start_km: float, end_km: float,
                             direction=None, lane=None):
        """모식도에서 지정 노선의 구간으로 이동하고 2.5초간 강조 표시합니다."""
        target_idx = next((i for i, r in enumerate(self.routes) if r["name"] == route_name), -1)
        if target_idx < 0:
            self._show_error("노선 없음", f"'{route_name}' 노선을 찾을 수 없습니다.")
            return

        # 탭 전환
        self.current_route_index = target_idx
        try:
            self.notebook.set(route_name)
        except Exception:
            pass
        self.refresh_route_controls_from_route(self.routes[target_idx])
        self.draw_schematic()

        # 렌더링 후 스크롤 + 강조
        self.after(180, lambda: self._scroll_and_highlight(target_idx, start_km, end_km,
                                                           direction, lane))

    def _scroll_and_highlight(self, route_idx: int, start_km: float, end_km: float,
                               direction=None, lane=None):
        """캔버스를 해당 구간으로 스크롤하고 노란 테두리로 강조합니다."""
        if not (0 <= route_idx < len(self.routes)):
            return
        route  = self.routes[route_idx]
        canvas = route.get("canvas")
        if canvas is None:
            return

        r_start  = float(route["start_km"])
        r_end    = float(route["end_km"])
        span_km  = max(0.1, r_end - r_start)
        margin   = 20

        # IC 갭을 고려한 x 위치 계산
        try:
            ic_gap_w    = IC_BOX_W + IC_GAP_MARGIN * 2
            ic_list     = self._collect_ic_list(route, r_start, r_end)
            total_gaps  = ic_gap_w * len(ic_list)

            # x_of_km: IC 삽입 누적 오프셋 반영
            def _x(km):
                offset = 0
                for ic_km, _ in ic_list:
                    if km > ic_km:
                        offset += ic_gap_w
                    else:
                        break
                return margin + (km - r_start) * PX_PER_KM + offset

            total_w = int(span_km * PX_PER_KM) + 80 + total_gaps
        except Exception:
            def _x(km):
                return margin + (km - r_start) * PX_PER_KM
            total_w = int(span_km * PX_PER_KM) + 80

        x1 = _x(start_km)
        x2 = _x(end_km)

        # 캔버스 뷰포트 폭
        canvas.update_idletasks()
        view_w = canvas.winfo_width() or 800

        # 구간 중심이 뷰포트 중앙에 오도록 스크롤
        center_x  = (x1 + x2) / 2
        scroll_x  = max(0.0, min(1.0, (center_x - view_w / 2) / max(total_w, 1)))
        canvas.xview_moveto(scroll_x)

        # 방향/차로에 따른 y 범위 계산
        h = canvas.winfo_height() or 360
        bl = route.get("_bar_layout", {})
        if bl and direction:
            bar1_top   = bl.get("bar1_top", 60)
            bar2_top   = bl.get("bar2_top", 200)
            bar_h      = bl.get("bar_h", 120)
            lane_count = route.get("lane_count", 4)
            dirs       = route.get("directions", [])
            def _norm(s): return s.replace("방향", "").strip()
            top = bar1_top if (not dirs or _norm(direction) == _norm(dirs[0])) else bar2_top

            if not lane or lane == "전차로":
                y1_hl = top - 2
                y2_hl = top + bar_h + 2
            else:
                try:
                    n = int(lane.replace("차로", "").strip())
                except (ValueError, AttributeError):
                    n = 1
                n = max(1, min(n, lane_count))
                seg_h = bar_h / lane_count
                if top == bar1_top:
                    y1_hl = top + bar_h - n * seg_h
                    y2_hl = top + bar_h - (n - 1) * seg_h
                else:
                    y1_hl = top + (n - 1) * seg_h
                    y2_hl = top + n * seg_h
        else:
            y1_hl = 4
            y2_hl = h - 4

        # 강조 사각형
        canvas.delete("nav_highlight")
        canvas.create_rectangle(x1, y1_hl, x2, y2_hl,
                                 outline="#FBBF24", width=3,
                                 fill="#FBBF24", stipple="gray25",
                                 tags=("nav_highlight",))
        canvas.create_line(x1, y1_hl, x1, y2_hl, fill="#FBBF24", width=2,
                            dash=(6, 4), tags=("nav_highlight",))
        canvas.create_line(x2, y1_hl, x2, y2_hl, fill="#FBBF24", width=2,
                            dash=(6, 4), tags=("nav_highlight",))
        canvas.tag_raise("nav_highlight")

        # 2.5초 후 강조 제거
        self.after(2500, lambda: canvas.delete("nav_highlight"))

